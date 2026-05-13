"""
Offline tests for the audit wizard.

These run without network — synthetic HTML pages are fed straight into the
Analyzer. The Crawler isn't tested live (network); we only verify it
constructs without errors.
"""

import io
import pytest

from audit_engine import (
    Analyzer, Crawler, ReportGen, detect_industry,
    INDUSTRY_PROFILES, ASTM_GRADES, MECHANICAL_PROPS, EQUIVALENT_GRADES,
    GRADE_RE, score_color,
)
import jobs


# ============================================================
#  Helpers
# ============================================================

def make_page(url: str, html: str, response_time: float = 0.5,
              content_length: int = 1000, status: int = 200) -> dict:
    """Build a fake page dict shaped like Crawler.fetch() returns."""
    return {
        "url": url,
        "html": html,
        "response_time": response_time,
        "content_length": content_length or len(html),
        "status": status,
    }


SAMPLE_METALS_HTML = """\
<html>
<head>
  <title>SS 304 Seamless Pipe Manufacturer | Example Co</title>
  <meta name="description" content="Premium SS 304 stainless steel seamless pipes manufactured to ASTM A312 standards. Various sizes for industrial use worldwide.">
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <link rel="canonical" href="https://example.com/304-pipe">
  <link rel="icon" href="/favicon.ico">
  <script type="application/ld+json">{"@type":"Product"}</script>
  <meta property="og:title" content="SS 304 Pipe">
</head>
<body>
  <h1>SS 304 Seamless Pipe</h1>
  <h2>Chemical Composition</h2>
  <p>Carbon (C): 0.08 max, Manganese (Mn): 2.00, Chromium (Cr): 18-20%, Nickel (Ni): 8-10.5%</p>
  <h2>Mechanical Properties</h2>
  <p>Tensile Strength: 515 MPa min, Yield Strength: 205 MPa min, Elongation: 40% min, Hardness: 201 HB max</p>
  <h2>Standards</h2>
  <p>Manufactured to ASTM A312 / ASME SA312. Equivalent grades: UNS S30400, EN 1.4301, DIN X5CrNi18-10, JIS SUS304.</p>
  <h2>Applications</h2>
  <p>Used in chemical plants, dairy equipment, and food processing industries.</p>
  <h2>Dimensions</h2>
  <p>Available in OD 6mm to 600mm, wall thickness 1mm to 60mm.</p>
  <h2>Contact</h2>
  <p>Phone: +91-9999999999, Email: sales@example.com. Get a quote today!</p>
  <img src="pipe1.jpg" alt="SS 304 pipe close-up">
  <img src="pipe2.jpg" alt="Pipe stack">
</body></html>
"""


# ============================================================
#  detect_industry
# ============================================================

def test_detect_industry_metals():
    text = "stainless steel pipe ASTM A312 grade 304 chromium"
    assert detect_industry("https://x.com/pipe", text) == "metals"


def test_detect_industry_ecommerce():
    text = "buy now add to cart checkout shipping price"
    assert detect_industry("https://shop.com/product", text) == "ecommerce"


def test_detect_industry_generic_when_unmatched():
    text = "lorem ipsum dolor sit amet"
    assert detect_industry("https://x.com", text) == "generic"


# ============================================================
#  Analyzer — universal checks
# ============================================================

def test_analyzer_returns_overall_score():
    page = make_page("https://example.com", SAMPLE_METALS_HTML)
    a = Analyzer(industry="metals")
    r = a.analyze(page)
    assert "overall" in r["scores"]
    assert 0 <= r["scores"]["overall"] <= 100


def test_analyzer_detects_metals_industry_with_auto():
    page = make_page("https://example.com/pipe", SAMPLE_METALS_HTML)
    a = Analyzer(industry="auto")
    r = a.analyze(page)
    assert r["industry"] == "metals"


def test_analyzer_returns_required_fields():
    page = make_page("https://example.com", SAMPLE_METALS_HTML)
    a = Analyzer()
    r = a.analyze(page)
    for field in ["url", "title", "issues", "scores", "word_count",
                  "grades_found", "response_time", "industry"]:
        assert field in r, f"missing field: {field}"


def test_analyzer_handles_empty_html():
    page = make_page("https://example.com", "", status=403)
    a = Analyzer()
    r = a.analyze(page)
    assert r["title"] == "[blocked]"
    assert r["issues"] == []


def test_analyzer_flags_missing_title():
    html = "<html><body><h1>Hello</h1></body></html>"
    a = Analyzer(industry="generic")
    r = a.analyze(make_page("https://x.com", html))
    titles = [i["title"] for i in r["issues"]]
    assert any("Title tag missing" in t for t in titles)


def test_analyzer_flags_missing_h1():
    html = "<html><head><title>OK title</title></head><body><p>no h1</p></body></html>"
    a = Analyzer(industry="generic")
    r = a.analyze(make_page("https://x.com", html))
    titles = [i["title"] for i in r["issues"]]
    assert any("H1 missing" in t for t in titles)


def test_analyzer_flags_noindex_as_critical():
    html = """<html><head><title>X</title><meta name="robots" content="noindex">
              </head><body><h1>X</h1></body></html>"""
    a = Analyzer(industry="generic")
    r = a.analyze(make_page("https://x.com", html))
    noindex_issues = [i for i in r["issues"] if "NOINDEX" in i["title"]]
    assert len(noindex_issues) == 1
    assert noindex_issues[0]["severity"] == "critical"


def test_analyzer_flags_http_as_critical():
    html = "<html><head><title>X</title></head><body><h1>x</h1></body></html>"
    a = Analyzer(industry="generic")
    r = a.analyze(make_page("http://insecure.com", html))
    http_issues = [i for i in r["issues"] if "HTTPS" in i["title"]]
    assert len(http_issues) >= 1


def test_analyzer_flags_slow_response_time():
    html = "<html><head><title>OK</title></head><body><h1>Slow</h1></body></html>"
    page = make_page("https://x.com", html, response_time=4.5)
    a = Analyzer(industry="generic")
    r = a.analyze(page)
    perf_issues = [i for i in r["issues"] if i["category"] == "performance"]
    assert any("Slow" in i["title"] for i in perf_issues)


# ============================================================
#  Analyzer — metals-specific checks
# ============================================================

def test_metals_check_misses_when_chemical_table_absent():
    html = """<html><head><title>SS 304 Pipe Manufacturer ASTM</title></head>
              <body><h1>SS 304 Pipe</h1><p>We make 304 pipes</p></body></html>"""
    a = Analyzer(industry="metals")
    r = a.analyze(make_page("https://x.com/304-pipe", html))
    cats = [i["category"] for i in r["issues"]]
    assert "chemical" in cats
    assert "mechanical" in cats


def test_metals_check_finds_grade_via_regex():
    matches = GRADE_RE.findall("Our SS 316L tubes meet ASTM A213 specifications")
    assert any("316L" in m or "316" in m for m in matches)
    assert any("A213" in m for m in matches)


# ============================================================
#  ReportGen
# ============================================================

@pytest.fixture
def sample_results():
    a = Analyzer(industry="auto")
    return [
        a.analyze(make_page("https://example.com/", SAMPLE_METALS_HTML)),
        a.analyze(make_page("https://example.com/about",
                            "<html><head><title>About Example Co</title></head>"
                            "<body><h1>About</h1><p>" + ("text " * 100) + "</p></body></html>")),
    ]


def test_reportgen_html_returns_complete_document(sample_results):
    gen = ReportGen()
    html = gen.html(sample_results, "https://example.com")
    assert html.startswith("<!DOCTYPE html>")
    assert "</html>" in html
    assert "Website Audit Report" in html
    assert "https://example.com/" in html


def test_reportgen_excel_returns_valid_xlsx_bytes(sample_results):
    gen = ReportGen()
    blob = gen.excel(sample_results)
    assert isinstance(blob, bytes)
    # XLSX files are zip files — start with PK
    assert blob[:2] == b"PK"
    # Verify openpyxl can read it back
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    assert "Summary" in wb.sheetnames
    assert "All Issues" in wb.sheetnames
    assert "Chemical Reference" in wb.sheetnames
    assert "Equivalent Grades" in wb.sheetnames


def test_reportgen_csv_has_header_and_rows(sample_results):
    gen = ReportGen()
    csv_text = gen.csv_report(sample_results)
    lines = csv_text.strip().split("\n")
    assert lines[0].startswith("URL,Title")
    assert len(lines) >= len(sample_results) + 1  # header + rows


# ============================================================
#  Score color
# ============================================================

def test_score_color_returns_green_for_high_score():
    fg, bg = score_color(85)
    assert fg.startswith("#22")  # dark green


def test_score_color_returns_red_for_low_score():
    fg, bg = score_color(40)
    assert fg.startswith("#74")  # dark red


# ============================================================
#  Reference data sanity
# ============================================================

def test_astm_grades_database_has_common_grades():
    for g in ["304", "304L", "316", "316L", "2205"]:
        assert g in ASTM_GRADES


def test_mechanical_props_match_grades():
    # Every grade with mechanical props should also have chemical data
    for g in MECHANICAL_PROPS:
        assert g in ASTM_GRADES, f"Grade {g} has mechanical but no chemical data"


def test_equivalent_grades_have_uns():
    for g, eq in EQUIVALENT_GRADES.items():
        assert "UNS" in eq


def test_industry_profiles_all_have_required_keys():
    for k, p in INDUSTRY_PROFILES.items():
        assert "name" in p
        assert "keywords" in p
        assert "required_sections" in p


# ============================================================
#  Job manager
# ============================================================

def test_job_create_assigns_id_and_starts_pending():
    j = jobs.create_job("https://x.com", "auto", 30)
    assert j.id
    assert j.status == "pending"
    assert j.url == "https://x.com"
    assert j.max_pages == 30


def test_job_get_returns_created_job():
    j = jobs.create_job("https://y.com", "auto", 10)
    same = jobs.get_job(j.id)
    assert same is j


def test_job_get_unknown_id_returns_none():
    assert jobs.get_job("does-not-exist") is None


def test_job_progress_updates():
    j = jobs.create_job("https://z.com", "auto", 50)
    j.update_progress(7, 50, "https://z.com/page-7")
    s = j.to_status_dict()
    assert s["current"] == 7
    assert s["total"] == 50
    assert s["current_url"] == "https://z.com/page-7"
    assert s["percent"] == 14


def test_job_cancel_signals_stop():
    j = jobs.create_job("https://c.com", "auto", 5)
    assert not j.should_stop()
    j.cancel()
    assert j.should_stop()


def test_get_summary_stats_handles_empty_results():
    j = jobs.create_job("https://x.com", "auto", 10)
    stats = jobs.get_summary_stats(j)
    assert stats["page_count"] == 0
    assert stats["avg_score"] == 0


# ============================================================
#  Crawler construction (no network)
# ============================================================

def test_crawler_construction_does_not_crash():
    c = Crawler("https://example.com", max_pages=10)
    assert c.base_url == "https://example.com"
    assert c.base_domain == "example.com"
    assert c.max_pages == 10


def test_crawler_threads_clamped_to_safe_range():
    c1 = Crawler("https://x.com", threads=0)
    c2 = Crawler("https://x.com", threads=99)
    assert c1.threads == 1
    assert c2.threads == 10


def test_crawler_extract_links_filters_other_domains():
    c = Crawler("https://example.com", max_pages=10)
    html = """<html><body>
        <a href="/about">About</a>
        <a href="https://example.com/contact">Contact</a>
        <a href="https://other-site.com/page">Other</a>
        <a href="mailto:hi@example.com">Email</a>
        <a href="file.pdf">PDF (skip)</a>
    </body></html>"""
    links = c.extract_links(html, "https://example.com/")
    # Only same-domain http(s) links, no PDFs
    assert "https://example.com/about" in links
    assert "https://example.com/contact" in links
    assert not any("other-site.com" in u for u in links)
    assert not any(".pdf" in u for u in links)


# ============================================================
#  Flask routes
# ============================================================

@pytest.fixture
def client():
    from app import app
    app.config["TESTING"] = True
    with app.test_client() as c:
        yield c


def test_root_redirects_to_step1(client):
    r = client.get("/")
    assert r.status_code == 302
    assert "/step1" in r.location


def test_step1_renders_form(client):
    r = client.get("/step1")
    assert r.status_code == 200
    assert b"<form" in r.data
    assert b"Website URL" in r.data


def test_healthz(client):
    r = client.get("/healthz")
    assert r.status_code == 200
    assert r.get_json() == {"status": "ok"}


def test_step1_post_creates_job_and_redirects(client):
    r = client.post(
        "/step1",
        data={"url": "https://example.com", "industry": "auto",
              "max_pages": "5", "email": ""},
        follow_redirects=False,
    )
    assert r.status_code == 302
    assert "/step2/" in r.location


def test_step1_post_rejects_invalid_email(client):
    r = client.post(
        "/step1",
        data={"url": "https://example.com", "industry": "auto",
              "max_pages": "5", "email": "not-an-email"},
        follow_redirects=False,
    )
    # Re-renders the form, doesn't redirect
    assert r.status_code == 200
    assert b"invalid" in r.data.lower()


def test_unknown_job_status_returns_404(client):
    r = client.get("/api/job/unknown_id_123/status")
    assert r.status_code == 404


def test_unknown_download_returns_404(client):
    r = client.get("/job/unknown_id_123/download.html")
    assert r.status_code == 404


# ============================================================
# Tests for the 27-point checklist build (stage 10)
# ============================================================

from audit_engine import (
    CHECKLIST, CHECKLIST_BY_ID, MANUAL_ONLY_CHECKPOINTS,
    build_checklist_view, aggregate_checklist,
)
from bs4 import BeautifulSoup


def test_checklist_has_exactly_27_items():
    assert len(CHECKLIST) == 27
    nums = sorted(c['number'] for c in CHECKLIST)
    assert nums == list(range(1, 28))


def test_checklist_ids_are_unique():
    ids = [c['id'] for c in CHECKLIST]
    assert len(ids) == len(set(ids))


def test_checklist_lookup_dict_works():
    assert 'page_title' in CHECKLIST_BY_ID
    assert CHECKLIST_BY_ID['page_title']['number'] == 2


def test_manual_only_set_contains_content_unique():
    assert 'content_unique' in MANUAL_ONLY_CHECKPOINTS


def test_url_structure_flags_query_params():
    a = Analyzer()
    issues = a._check_url_structure('https://example.com/product?id=12345')
    assert any('query parameters' in i['title'].lower() for i in issues)


def test_url_structure_flags_underscores():
    a = Analyzer()
    issues = a._check_url_structure('https://example.com/ss_304_pipe.html')
    assert any('underscore' in i['title'].lower() for i in issues)


def test_url_structure_flags_uppercase():
    a = Analyzer()
    issues = a._check_url_structure('https://example.com/Products/SS304')
    assert any('uppercase' in i['title'].lower() for i in issues)


def test_url_structure_clean_url_no_issues():
    a = Analyzer()
    issues = a._check_url_structure('https://example.com/products/ss-304-pipe')
    assert issues == []


def test_breadcrumbs_detected_via_schema():
    a = Analyzer()
    html = '''<html><body>
    <script type="application/ld+json">{"@type":"BreadcrumbList","itemListElement":[]}</script>
    </body></html>'''
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_breadcrumbs(soup)
    # Has schema, no visible — gets info note about visible breadcrumbs missing
    # but should NOT get the "missing" warning
    assert not any('Breadcrumbs missing' in i['title'] for i in issues)


def test_breadcrumbs_missing_flagged():
    a = Analyzer()
    html = '<html><body><h1>Page</h1></body></html>'
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_breadcrumbs(soup)
    assert any('Breadcrumbs missing' in i['title'] for i in issues)


def test_image_filenames_flags_generic():
    a = Analyzer()
    html = '<html><body><img src="IMG_1234.jpg"><img src="DSC_5678.jpg"></body></html>'
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_image_filenames(soup)
    assert len(issues) == 1
    assert '2 images' in issues[0]['title'] or 'generic filenames' in issues[0]['title']


def test_image_filenames_descriptive_no_issues():
    a = Analyzer()
    html = '<html><body><img src="ss-304-pipe.jpg"><img src="chemical-table.png"></body></html>'
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_image_filenames(soup)
    assert issues == []


def test_internal_linking_counts_links_correctly():
    a = Analyzer()
    html = '''<html><body>
    <a href="/product/ss-304">SS 304 Pipe Details</a>
    <a href="/product/ss-316">SS 316 Pipe Details</a>
    <a href="/product/duplex">Duplex Steel Pipe</a>
    </body></html>'''
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_internal_linking(soup, 'https://example.com/')
    # Should have 3 substantive internal links — no warning
    assert issues == []


def test_internal_linking_warns_when_few_links():
    a = Analyzer()
    html = '<html><body><a href="#">→</a></body></html>'
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_internal_linking(soup, 'https://example.com/')
    assert any('few internal links' in i['title'].lower() for i in issues)


def test_keyword_usage_skipped_when_no_target():
    a = Analyzer()
    issues = a._check_keyword_usage(BeautifulSoup('<html></html>', 'lxml'), 'some text', '')
    assert issues == []


def test_keyword_usage_flags_missing_keyword():
    a = Analyzer()
    html = '<html><head><title>Generic Page</title></head><body><h1>Generic</h1><p>Some unrelated content text.</p></body></html>'
    soup = BeautifulSoup(html, 'lxml')
    text = soup.get_text(' ', strip=True)
    issues = a._check_keyword_usage(soup, text, 'stainless steel')
    # The check raises a critical issue with "not in page content" in the title
    assert any('not in page content' in i['title'].lower() for i in issues)


def test_keyword_usage_flags_overuse():
    a = Analyzer()
    html = '<html><body><h1>Pipe</h1></body></html>'
    soup = BeautifulSoup(html, 'lxml')
    # Density > 3% — make 5 of 10 words be the keyword
    text = "pipe pipe pipe pipe pipe other other other other other"
    issues = a._check_keyword_usage(soup, text, 'pipe')
    assert any('over-used' in i['title'].lower() for i in issues)


def test_specifications_table_flags_missing():
    a = Analyzer()
    html = '<html><body><h1>Page</h1><p>No tables here.</p></body></html>'
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_product_tables(soup, soup.get_text())
    assert any('No tables on page' in i['title'] for i in issues)


def test_applications_section_present_passes():
    a = Analyzer()
    html = '''<html><body>
    <h2>Applications</h2>
    <ul><li>Petrochem</li><li>Pharma</li><li>Food processing</li></ul>
    </body></html>'''
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_applications(soup, soup.get_text(), 'metals')
    assert issues == []


def test_applications_missing_flagged():
    a = Analyzer()
    html = '<html><body><h1>Page</h1></body></html>'
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_applications(soup, soup.get_text(), 'metals')
    assert any('No Applications' in i['title'] for i in issues)


def test_cta_buttons_dead_link_flagged():
    a = Analyzer()
    html = '<html><body><a href="#" class="btn">Get Quote</a></body></html>'
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_cta_buttons(soup, 'https://example.com/', deep=False)
    assert any('empty/dead link' in i['title'] for i in issues)


def test_cta_buttons_no_cta_flagged():
    a = Analyzer()
    html = '<html><body><h1>Page</h1></body></html>'
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_cta_buttons(soup, 'https://example.com/', deep=False)
    assert any('No CTA' in i['title'] for i in issues)


def test_faqs_detected_via_schema():
    a = Analyzer()
    html = '''<html><body>
    <script type="application/ld+json">{"@type":"FAQPage"}</script>
    </body></html>'''
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_faqs(soup, soup.get_text())
    # Schema present — no "missing" warning
    assert not any('No FAQ' in i['title'] for i in issues)


def test_faqs_via_details_element():
    a = Analyzer()
    html = '<html><body><details><summary>Q?</summary><p>A.</p></details></body></html>'
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_faqs(soup, soup.get_text())
    # Has accordion, no schema — info-level recommendation about schema
    assert not any(i['severity'] == 'critical' for i in issues)


def test_responsive_no_media_query_flagged():
    a = Analyzer()
    html = '<html><head><meta name="viewport" content="width=device-width, initial-scale=1"></head><body><p>x</p></body></html>'
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_responsive(soup, html)
    assert any('responsive CSS' in i['title'].lower() or 'responsive' in i['title'].lower() for i in issues)


def test_responsive_with_media_query_passes():
    a = Analyzer()
    html = '<html><head><style>@media (max-width:768px){.x{padding:1em}}</style></head><body></body></html>'
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_responsive(soup, html)
    # Should not flag missing responsive — but may flag missing viewport
    assert not any('No responsive CSS detected' in i['title'] for i in issues)


def test_inquiry_form_with_email_passes():
    a = Analyzer()
    html = '''<html><body><form action="/submit">
    <input name="name"><input name="email" type="email"><textarea name="msg"></textarea>
    <button type="submit">Send</button></form></body></html>'''
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_inquiry_form(soup)
    assert issues == []


def test_inquiry_form_missing_flagged():
    a = Analyzer()
    html = '<html><body><h1>No form here</h1></body></html>'
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_inquiry_form(soup)
    assert any('No inquiry form' in i['title'] for i in issues)


def test_schema_validation_flags_broken_json():
    a = Analyzer()
    html = '''<html><body>
    <script type="application/ld+json">{ "this is": "broken json", }</script>
    </body></html>'''
    soup = BeautifulSoup(html, 'lxml')
    issues = a._check_schema_validation(soup)
    assert any('parse errors' in i['title'].lower() for i in issues)


def test_build_checklist_view_returns_27_items():
    view = build_checklist_view(issues=[])
    assert len(view) == 27


def test_build_checklist_view_pass_when_no_issues():
    view = build_checklist_view(issues=[])
    # All non-manual non-deep items should be 'pass'
    for entry in view:
        cp = CHECKLIST_BY_ID[entry['id']]
        if cp['auto'] == 'manual':
            assert entry['status'] == 'manual'
        elif cp['auto'] == 'deep':
            assert entry['status'] == 'skipped'
        elif cp['id'] == 'keyword_usage':
            assert entry['status'] == 'skipped'  # no target keyword
        else:
            assert entry['status'] == 'pass'


def test_build_checklist_view_fail_on_critical_issue():
    issues = [{'category': 'seo', 'severity': 'critical', 'title': 'Title missing',
               'description': '', 'fix': '', 'checkpoint_id': 'page_title'}]
    view = build_checklist_view(issues)
    title_entry = next(e for e in view if e['id'] == 'page_title')
    assert title_entry['status'] == 'fail'
    assert title_entry['severity'] == 'critical'
    assert title_entry['issue_count'] == 1


def test_build_checklist_view_deep_skipped_unless_enabled():
    view_off = build_checklist_view(issues=[], deep_audit=False)
    view_on = build_checklist_view(issues=[], deep_audit=True)
    img_off = next(e for e in view_off if e['id'] == 'image_weight')
    img_on = next(e for e in view_on if e['id'] == 'image_weight')
    assert img_off['status'] == 'skipped'
    assert img_on['status'] == 'pass'


def test_aggregate_checklist_counts_correctly():
    # Two pages: one passes everything, the other fails page_title
    pages = [
        {'checklist': build_checklist_view(issues=[])},
        {'checklist': build_checklist_view(
            issues=[{'category': 'seo', 'severity': 'critical', 'title': 'X',
                     'description': '', 'fix': '', 'checkpoint_id': 'page_title'}]
        )},
    ]
    agg = aggregate_checklist(pages)
    title_agg = next(a for a in agg if a['id'] == 'page_title')
    assert title_agg['pass'] == 1
    assert title_agg['fail'] == 1


def test_analyzer_accepts_target_keyword_and_deep_audit():
    a = Analyzer(industry='auto', target_keyword='ss 304 pipe', deep_audit=True)
    assert a.target_keyword == 'ss 304 pipe'
    assert a.deep_audit is True


def test_analyzer_returns_checklist_in_result():
    a = Analyzer()
    page = {
        'url': 'https://example.com/',
        'html': '<html><head><title>Test Page</title></head><body><h1>Test</h1></body></html>',
        'response_time': 0.5, 'content_length': 200, 'status': 200,
    }
    result = a.analyze(page)
    assert 'checklist' in result
    assert len(result['checklist']) == 27


def test_excel_report_contains_checklist_sheet():
    rg = ReportGen()
    a = Analyzer()
    page = {
        'url': 'https://example.com/',
        'html': '<html><head><title>Test</title></head><body><h1>X</h1></body></html>',
        'response_time': 0.3, 'content_length': 100, 'status': 200,
    }
    result = a.analyze(page)
    xlsx_bytes = rg.excel([result])

    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert '27-Point Checklist' in wb.sheetnames


def test_html_report_contains_checklist_block():
    rg = ReportGen()
    a = Analyzer()
    page = {
        'url': 'https://example.com/',
        'html': '<html><head><title>Test</title></head><body><h1>X</h1></body></html>',
        'response_time': 0.3, 'content_length': 100, 'status': 200,
    }
    result = a.analyze(page)
    html = rg.html([result], 'https://example.com/')
    assert '27-Point Checklist Summary' in html
    assert 'class="cls"' in html


# ============================================================
# Tests for the streaming crawler + recommendations + new checks
# ============================================================

from unittest.mock import patch, MagicMock
from audit_engine import (
    generate_recommendations, top_recommendations, overall_health_summary,
    RECOMMENDATION_THEMES,
)
from bs4 import BeautifulSoup as BS


def _fake_audit_result(url, critical_checkpoints=None, warning_checkpoints=None):
    """Helper: build a minimal audit result dict for testing."""
    issues = []
    for cp in (critical_checkpoints or []):
        issues.append({'category': 'seo', 'severity': 'critical', 'title': 'x',
                       'description': '', 'fix': '', 'checkpoint_id': cp})
    for cp in (warning_checkpoints or []):
        issues.append({'category': 'seo', 'severity': 'warning', 'title': 'x',
                       'description': '', 'fix': '', 'checkpoint_id': cp})
    return {
        'url': url, 'title': 'X', 'issues': issues,
        'scores': {'overall': 50 if issues else 100},
        'word_count': 200, 'grades_found': [],
        'response_time': 1, 'industry': 'metals',
        'status': 200, 'checklist': [],
    }


def test_generate_recommendations_empty_results():
    assert generate_recommendations([]) == []


def test_generate_recommendations_sorts_by_priority():
    # Page 1: critical h1 + critical title; Page 2: critical h1 + warning alt
    # h1 affects both pages, title affects 1 page critical, alt affects 1 page warning
    results = [
        _fake_audit_result('p1', critical_checkpoints=['h1_tag', 'page_title']),
        _fake_audit_result('p2', critical_checkpoints=['h1_tag'],
                           warning_checkpoints=['image_alt']),
    ]
    recs = generate_recommendations(results)
    assert len(recs) >= 3

    # h1_tag should be first (critical + affects ALL pages)
    assert recs[0]['id'] == 'h1_tag'
    assert recs[0]['severity'] == 'critical'
    assert recs[0]['affected_pages'] == 2

    # Later ones should have lower priority scores
    for i in range(1, len(recs)):
        assert recs[i]['priority_score'] <= recs[i-1]['priority_score']


def test_top_recommendations_limits_count():
    results = [_fake_audit_result(f'p{i}',
                                  critical_checkpoints=['h1_tag', 'page_title',
                                                        'canonical', 'image_alt',
                                                        'schema', 'breadcrumbs'])
               for i in range(3)]
    recs = top_recommendations(results, n=3)
    assert len(recs) == 3


def test_recommendation_has_required_fields():
    results = [_fake_audit_result('p1', critical_checkpoints=['h1_tag'])]
    recs = generate_recommendations(results)
    rec = recs[0]
    assert 'theme' in rec
    assert 'theme_label' in rec
    assert 'headline' in rec
    assert 'why' in rec
    assert 'action_steps' in rec
    assert isinstance(rec['action_steps'], list)
    assert len(rec['action_steps']) > 0
    assert rec['affected_pages'] == 1
    assert rec['severity'] == 'critical'


def test_recommendation_themes_are_valid():
    results = [_fake_audit_result('p1', critical_checkpoints=['h1_tag', 'image_alt'])]
    for rec in generate_recommendations(results):
        assert rec['theme'] in RECOMMENDATION_THEMES, \
            f"Unknown theme: {rec['theme']}"


def test_overall_health_summary_computes_correctly():
    results = [
        _fake_audit_result('p1', critical_checkpoints=['h1_tag']),
        _fake_audit_result('p2'),  # clean page
        _fake_audit_result('p3', warning_checkpoints=['image_alt']),
    ]
    h = overall_health_summary(results)
    assert h['page_count'] == 3
    assert h['critical_issues'] == 1
    assert h['warning_issues'] == 1
    assert h['pages_with_critical'] == 1
    assert h['grade'] in ('A', 'B', 'C', 'D', 'F')


def test_overall_health_summary_empty_results():
    h = overall_health_summary([])
    assert h['page_count'] == 0
    assert h['avg_score'] == 0
    assert h['grade'] == 'N/A'


def test_security_headers_check_flags_missing_hsts():
    a = Analyzer()
    pg = {'response_headers': {}}
    issues = a._check_security_headers(pg)
    titles = [i['title'] for i in issues]
    assert any('HSTS' in t for t in titles)
    assert any('Content-Security-Policy' in t for t in titles)


def test_security_headers_check_silent_when_no_headers():
    """If we couldn't get headers at all, don't spam issues."""
    a = Analyzer()
    issues = a._check_security_headers({})  # No response_headers key
    assert issues == []


def test_mixed_content_flags_http_resources_on_https_page():
    a = Analyzer()
    html = '''<html><body>
    <img src="http://example.com/logo.png">
    <script src="http://example.com/analytics.js"></script>
    </body></html>'''
    soup = BS(html, 'lxml')
    issues = a._check_mixed_content(soup, 'https://example.com/page')
    assert len(issues) == 1
    assert '2 resources' in issues[0]['title']


def test_mixed_content_skipped_on_http_page():
    a = Analyzer()
    html = '<html><body><img src="http://x.com/y.png"></body></html>'
    soup = BS(html, 'lxml')
    issues = a._check_mixed_content(soup, 'http://example.com/page')
    assert issues == []  # No HTTPS, no mixed-content concern


def test_social_media_tags_flags_missing_og_image():
    a = Analyzer()
    soup = BS('<html><head></head><body></body></html>', 'lxml')
    issues = a._check_social_media_tags(soup)
    titles = [i['title'] for i in issues]
    assert any('og:image' in t for t in titles)
    assert any('Twitter Card' in t for t in titles)


def test_social_media_tags_passes_when_present():
    a = Analyzer()
    html = '''<html><head>
    <meta property="og:image" content="https://x.com/og.jpg">
    <meta name="twitter:card" content="summary_large_image">
    </head></html>'''
    soup = BS(html, 'lxml')
    issues = a._check_social_media_tags(soup)
    assert issues == []


def test_streaming_crawler_yields_pages_one_at_a_time():
    """Verify crawl_streaming() is a generator that yields incrementally."""
    crawler = Crawler('https://example.com', max_pages=3)
    # The method must be a generator (callable returning iterator)
    import inspect
    assert inspect.isgeneratorfunction(crawler.crawl_streaming)


def test_seed_queue_from_sitemap_uses_sitemap_urls():
    """When sitemap exists, the queue is reseeded with its URLs."""
    crawler = Crawler('https://example.com', max_pages=10)
    fake_sitemap_xml = '''<?xml version="1.0"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
  <url><loc>https://example.com/p1</loc></url>
  <url><loc>https://example.com/p2</loc></url>
</urlset>'''

    def fake_get(url, *a, **kw):
        m = MagicMock()
        if 'robots.txt' in url:
            m.status_code = 404
            m.text = ''
        elif 'sitemap.xml' in url:
            m.status_code = 200
            m.text = fake_sitemap_xml
        else:
            m.status_code = 404
            m.text = ''
        return m

    with patch.object(crawler.session, 'get', side_effect=fake_get):
        count = crawler._seed_queue_from_sitemap()

    assert count == 2
    # Queue should now have those URLs
    queued = []
    while not crawler.q.empty():
        queued.append(crawler.q.get_nowait())
    assert 'https://example.com/p1' in queued
    assert 'https://example.com/p2' in queued
