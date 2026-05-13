"""
Audit engine — the heart of the wizard.

Refactored from website_auditor_v2's audit_agent.py. Same logic, but reorganised
for web use: the Crawler and Analyzer accept progress callbacks so the Flask app
can stream live progress to the browser, and the report generators take a
results list and return strings/bytes (no direct file writing).

Public API:
    Crawler(base_url, max_pages, threads, delay, timeout, skip_ssl, on_progress, should_stop)
        .crawl() -> list[dict]
    Analyzer(industry='auto')
        .analyze(page) -> dict
    ReportGen()
        .html(results, site_url) -> str
        .excel(results) -> bytes
        .csv_report(results) -> str
    detect_industry(url, text) -> str
"""

from __future__ import annotations

import csv
import io
import re
import threading
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from queue import Queue
from typing import Callable, Optional
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill
from requests.packages.urllib3.exceptions import InsecureRequestWarning

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)


# ============================================================
#  Industry profiles
# ============================================================

INDUSTRY_PROFILES = {
    "metals": {
        "name": "Metals / Steel / Industrial",
        "keywords": ["steel", "pipe", "tube", "plate", "bar", "flange", "fitting",
                     "alloy", "grade", "astm", "asme", "ss", "stainless"],
        "required_sections": ["chemical composition", "mechanical properties",
                              "dimensions", "equivalent grades", "applications", "standards"],
    },
    "ecommerce": {
        "name": "E-Commerce / Online Store",
        "keywords": ["buy", "cart", "checkout", "product", "price", "shop", "order", "shipping"],
        "required_sections": ["price", "description", "specifications", "reviews", "shipping"],
    },
    "saas": {
        "name": "SaaS / Software",
        "keywords": ["software", "platform", "dashboard", "api", "pricing", "free trial", "demo"],
        "required_sections": ["features", "pricing", "testimonials", "faq"],
    },
    "healthcare": {
        "name": "Healthcare / Medical",
        "keywords": ["doctor", "hospital", "clinic", "patient", "treatment", "medical", "appointment"],
        "required_sections": ["services", "doctors", "appointment", "contact"],
    },
    "realestate": {
        "name": "Real Estate",
        "keywords": ["property", "flat", "apartment", "villa", "bhk", "sqft", "rent", "buy"],
        "required_sections": ["price", "location", "amenities", "contact"],
    },
    "generic": {
        "name": "Generic Website",
        "keywords": [],
        "required_sections": ["about", "contact", "services"],
    },
}


# ============================================================
#  Metals reference data
# ============================================================

ASTM_GRADES = {
    "304":  {"C": "0.08", "Mn": "2.00", "Si": "0.75", "P": "0.045", "S": "0.030",
             "Cr": "18.0-20.0", "Ni": "8.0-10.5"},
    "304L": {"C": "0.030", "Mn": "2.00", "Si": "0.75", "P": "0.045", "S": "0.030",
             "Cr": "18.0-20.0", "Ni": "8.0-12.0"},
    "316":  {"C": "0.08", "Mn": "2.00", "Si": "0.75", "P": "0.045", "S": "0.030",
             "Cr": "16.0-18.0", "Ni": "10.0-14.0", "Mo": "2.0-3.0"},
    "316L": {"C": "0.030", "Mn": "2.00", "Si": "0.75", "P": "0.045", "S": "0.030",
             "Cr": "16.0-18.0", "Ni": "10.0-14.0", "Mo": "2.0-3.0"},
    "321":  {"C": "0.08", "Mn": "2.00", "Si": "0.75", "P": "0.045", "S": "0.030",
             "Cr": "17.0-19.0", "Ni": "9.0-12.0", "Ti": "5xC min"},
    "347":  {"C": "0.08", "Mn": "2.00", "Si": "0.75", "P": "0.045", "S": "0.030",
             "Cr": "17.0-19.0", "Ni": "9.0-13.0"},
    "310S": {"C": "0.08", "Mn": "2.00", "Si": "1.50", "P": "0.045", "S": "0.030",
             "Cr": "24.0-26.0", "Ni": "19.0-22.0"},
    "317L": {"C": "0.030", "Mn": "2.00", "Si": "0.75", "P": "0.045", "S": "0.030",
             "Cr": "18.0-20.0", "Ni": "11.0-15.0", "Mo": "3.0-4.0"},
    "904L": {"C": "0.020", "Mn": "2.00", "Si": "1.00", "P": "0.045", "S": "0.035",
             "Cr": "19.0-23.0", "Ni": "23.0-28.0", "Mo": "4.0-5.0"},
    "2205": {"C": "0.030", "Mn": "2.00", "Si": "1.00", "P": "0.030", "S": "0.020",
             "Cr": "22.0-23.0", "Ni": "4.5-6.5", "Mo": "3.0-3.5", "N": "0.14-0.20"},
    "2507": {"C": "0.030", "Mn": "1.20", "Si": "0.80", "P": "0.035", "S": "0.020",
             "Cr": "24.0-26.0", "Ni": "6.0-8.0", "Mo": "3.0-5.0"},
    "A36":  {"C": "0.26", "P": "0.04", "S": "0.05"},
    "A53":  {"C": "0.25", "Mn": "0.95", "P": "0.05", "S": "0.045"},
    "A106": {"C": "0.30", "Mn": "0.29-1.06", "Si": "0.10", "P": "0.035", "S": "0.035"},
}

MECHANICAL_PROPS = {
    "304":  {"tensile": "515 MPa min", "yield": "205 MPa min", "elongation": "40% min", "hardness": "201 HB max"},
    "304L": {"tensile": "485 MPa min", "yield": "170 MPa min", "elongation": "40% min", "hardness": "201 HB max"},
    "316":  {"tensile": "515 MPa min", "yield": "205 MPa min", "elongation": "40% min", "hardness": "217 HB max"},
    "316L": {"tensile": "485 MPa min", "yield": "170 MPa min", "elongation": "40% min", "hardness": "217 HB max"},
    "321":  {"tensile": "515 MPa min", "yield": "205 MPa min", "elongation": "40% min", "hardness": "217 HB max"},
    "310S": {"tensile": "515 MPa min", "yield": "205 MPa min", "elongation": "40% min", "hardness": "217 HB max"},
    "904L": {"tensile": "490 MPa min", "yield": "220 MPa min", "elongation": "35% min", "hardness": "—"},
    "2205": {"tensile": "620 MPa min", "yield": "450 MPa min", "elongation": "25% min", "hardness": "293 HB max"},
    "2507": {"tensile": "800 MPa min", "yield": "550 MPa min", "elongation": "15% min", "hardness": "310 HB max"},
}

EQUIVALENT_GRADES = {
    "304":  {"UNS": "S30400", "EN": "1.4301", "DIN": "X5CrNi18-10", "JIS": "SUS304", "BS": "304S31"},
    "304L": {"UNS": "S30403", "EN": "1.4307", "DIN": "X2CrNi19-11", "JIS": "SUS304L", "BS": "304S11"},
    "316":  {"UNS": "S31600", "EN": "1.4401", "DIN": "X5CrNiMo17-12-2", "JIS": "SUS316", "BS": "316S31"},
    "316L": {"UNS": "S31603", "EN": "1.4404", "DIN": "X2CrNiMo17-12-2", "JIS": "SUS316L", "BS": "316S11"},
    "321":  {"UNS": "S32100", "EN": "1.4541", "DIN": "X6CrNiTi18-10", "JIS": "SUS321", "BS": "321S31"},
    "310S": {"UNS": "S31008", "EN": "1.4845", "DIN": "X8CrNi25-21", "JIS": "SUS310S", "BS": "310S24"},
    "904L": {"UNS": "N08904", "EN": "1.4539", "DIN": "X1NiCrMoCu25-20-5", "JIS": "SUS890L", "BS": "904S13"},
    "2205": {"UNS": "S32205", "EN": "1.4462", "DIN": "X2CrNiMoN22-5-3", "JIS": "SUS329J3L", "BS": "—"},
    "2507": {"UNS": "S32750", "EN": "1.4410", "DIN": "X2CrNiMoN25-7-4", "JIS": "—", "BS": "—"},
}

ASTM_STANDARDS = {
    "pipe":     ["ASTM A312", "ASTM A790", "ASTM A358", "ASTM A409", "ASME SA312"],
    "tube":     ["ASTM A213", "ASTM A249", "ASTM A269", "ASTM A270", "ASTM A554", "ASME SA213"],
    "plate":    ["ASTM A240", "ASTM A480", "ASME SA240"],
    "bar":      ["ASTM A276", "ASTM A479", "ASTM A484", "ASME SA479"],
    "fitting":  ["ASTM A403", "ASTM A182", "ASTM A815", "ASME SA403"],
    "flange":   ["ASTM A182", "ASTM A240", "ASME SA182"],
    "sheet":    ["ASTM A240", "ASTM A167", "ASTM A480"],
    "fastener": ["ASTM A193", "ASTM A194", "ASTM F593", "ASME SA193"],
}

GRADE_RE = re.compile(
    r'\b(SS\s*304L?|SS\s*316L?|SS\s*321|SS\s*310S?|SS\s*317L?|SS\s*904L|'
    r'304L?|316L?|321|310S?|317L?|904L|2205|2507|347|'
    r'A312|A213|A240|A276|A403|A182|A358|A790|A106|A53|A36|'
    r'ASTM\s+[A-Z]\d+|ASME\s+SA\d+|EN\s+1\.\d{4}|JIS\s+SUS\d+)\b',
    re.IGNORECASE,
)

SKIP_EXT = {
    '.pdf', '.jpg', '.jpeg', '.png', '.gif', '.svg', '.webp', '.ico',
    '.css', '.js', '.zip', '.gz', '.mp4', '.mp3', '.woff', '.woff2', '.ttf', '.eot',
}

SEV_ORDER = {'critical': 0, 'warning': 1, 'info': 2}


# ============================================================
#  27-Point Checklist (master registry)
# ============================================================
#
# This is the canonical list every audit issue ties back to. Each entry has:
#   id           : stable short slug used to tag issues (issue['checkpoint_id'])
#   number       : 1-27, the position on the checklist
#   label        : what shows in reports
#   description  : the rule, in plain language
#   group        : groups for the report's checklist view
#   auto         : "full"   = fully auto-checked
#                  "partial"= partially auto-checked, some judgment needed
#                  "deep"   = only checked when deep_audit=True (slow, network-heavy)
#                  "manual" = cannot be auto-checked; surfaced as "needs review"
#
# When you add a new check method, give every issue it raises a checkpoint_id
# matching one of these IDs. The checklist view in the report uses these tags
# to compute pass/fail per checkpoint per page.

CHECKLIST = [
    # On-page SEO basics (1-8)
    {"id": "url_structure",     "number": 1,  "label": "URL Structure",
     "description": "Clean, readable, keyword-rich URL.",
     "group": "On-page SEO",     "auto": "full"},
    {"id": "page_title",        "number": 2,  "label": "Page Title",
     "description": "Unique title with main keyword (60-70 characters max).",
     "group": "On-page SEO",     "auto": "full"},
    {"id": "meta_description",  "number": 3,  "label": "Meta Description",
     "description": "Page meta description with keywords (max 160 characters).",
     "group": "On-page SEO",     "auto": "full"},
    {"id": "indexability",      "number": 4,  "label": "Index / Noindex Tag",
     "description": "Page is indexable; no noindex tag present.",
     "group": "On-page SEO",     "auto": "full"},
    {"id": "canonical",         "number": 5,  "label": "Canonical Tag",
     "description": "Canonical link present and points to the correct URL.",
     "group": "On-page SEO",     "auto": "full"},
    {"id": "h1_tag",            "number": 6,  "label": "H1 Tag",
     "description": "Exactly one H1 per page, containing the main keyword.",
     "group": "On-page SEO",     "auto": "full"},
    {"id": "breadcrumbs",       "number": 7,  "label": "Breadcrumbs",
     "description": "Proper breadcrumbs with BreadcrumbList schema.",
     "group": "On-page SEO",     "auto": "full"},
    {"id": "subheadings",       "number": 8,  "label": "Subheadings",
     "description": "Proper H2/H3 structure with related products/topics.",
     "group": "On-page SEO",     "auto": "full"},

    # Images (9-11)
    {"id": "image_filename",    "number": 9,  "label": "Image File Names",
     "description": "Descriptive filenames (e.g. stainless-steel-pipe.jpg, not IMG_1234.jpg).",
     "group": "Images",          "auto": "full"},
    {"id": "image_alt",         "number": 10, "label": "Image Alt Text",
     "description": "All images have descriptive alt text.",
     "group": "Images",          "auto": "full"},
    {"id": "image_weight",      "number": 11, "label": "Image Optimization",
     "description": "Each image ≤ 100 KB.",
     "group": "Images",          "auto": "deep"},

    # Linking & content (12-14)
    {"id": "internal_linking",  "number": 12, "label": "Internal Linking",
     "description": "Links to related products or category pages.",
     "group": "Content",         "auto": "full"},
    {"id": "keyword_usage",     "number": 13, "label": "Keyword Usage",
     "description": "Target keyword present and not overused (1-3% density).",
     "group": "Content",         "auto": "partial"},
    {"id": "content_unique",    "number": 14, "label": "Content Originality",
     "description": "Not copied from other sites or catalogs.",
     "group": "Content",         "auto": "manual"},

    # Tables (15-20) — bread-and-butter for industrial product pages
    {"id": "specifications",    "number": 15, "label": "Specifications Table",
     "description": "Table with sizes, grades, standards.",
     "group": "Product Tables",  "auto": "full"},
    {"id": "chemical_table",    "number": 16, "label": "Chemical Table",
     "description": "Exact chemical composition for the actual grade (e.g. SS304, SS316).",
     "group": "Product Tables",  "auto": "partial"},
    {"id": "mechanical_table",  "number": 17, "label": "Mechanical Properties",
     "description": "Accurate mechanical properties for the specific grade.",
     "group": "Product Tables",  "auto": "partial"},
    {"id": "equivalent_table",  "number": 18, "label": "Equivalent Grades",
     "description": "Correct equivalency table for the grade. No dummy/unrelated standards.",
     "group": "Product Tables",  "auto": "partial"},
    {"id": "size_table",        "number": 19, "label": "Size / Dimension Table",
     "description": "Exact size/dimension table from official reference or datasheet.",
     "group": "Product Tables",  "auto": "partial"},
    {"id": "tech_spec_tab",     "number": 20, "label": "Technical Specification Tab",
     "description": "Spec table from a valid source/datasheet for the exact product.",
     "group": "Product Tables",  "auto": "partial"},

    # User-facing content (21-24)
    {"id": "applications",      "number": 21, "label": "Applications",
     "description": "Industry-specific applications, not generic lists.",
     "group": "Page Content",    "auto": "partial"},
    {"id": "cta_button",        "number": 22, "label": "CTA Button",
     "description": "CTA button present with a valid working link.",
     "group": "Page Content",    "auto": "partial"},
    {"id": "faqs",              "number": 23, "label": "FAQs",
     "description": "FAQ section with clear questions and technically-accurate answers.",
     "group": "Page Content",    "auto": "partial"},
    {"id": "contact_info",      "number": 24, "label": "Contact Info",
     "description": "Complete and correct contact details.",
     "group": "Page Content",    "auto": "full"},

    # Technical (25-27)
    {"id": "schema",            "number": 25, "label": "Schema Markup",
     "description": "Schema.org JSON-LD present, valid, and matching expected types.",
     "group": "Technical",       "auto": "full"},
    {"id": "mobile_friendly",   "number": 26, "label": "Mobile Friendly",
     "description": "Viewport set + responsive CSS hints. Verify visually on real devices.",
     "group": "Technical",       "auto": "partial"},
    {"id": "inquiry_form",      "number": 27, "label": "Inquiry Form",
     "description": "Form with email field and valid action. Manually verify it actually delivers.",
     "group": "Technical",       "auto": "partial"},
]

# Quick lookup: id -> entry
CHECKLIST_BY_ID = {c["id"]: c for c in CHECKLIST}

# Categories that should never be reported as "Pass" without a real signal —
# i.e. if there are zero issues for this checkpoint, it's because we couldn't
# auto-check it, not because the page passed. The UI shows these as "Manual review".
MANUAL_ONLY_CHECKPOINTS = {c["id"] for c in CHECKLIST if c["auto"] == "manual"}


# ============================================================
#  Helpers
# ============================================================

def detect_industry(url: str, text: str) -> str:
    """Pick the industry profile whose keywords best match the page."""
    combined = (url + " " + text[:3000]).lower()
    scores = {
        k: sum(1 for kw in v["keywords"] if kw in combined)
        for k, v in INDUSTRY_PROFILES.items() if k != "generic"
    }
    best = max(scores, key=scores.get) if scores else "generic"
    return best if scores.get(best, 0) > 0 else "generic"


def _iss(category: str, severity: str, title: str, description: str,
         fix: str = "", checkpoint_id: str = "") -> dict:
    """
    Build an issue dict.

    `checkpoint_id` ties the issue back to one of the 27 CHECKLIST items.
    Pass an ID like "page_title", "h1_tag", "image_alt", etc.
    Empty string means "not associated with a specific checkpoint" (rare;
    a few legacy informational items don't map cleanly).
    """
    return {
        "category": category,
        "severity": severity,
        "title": title,
        "description": description,
        "fix": fix,
        "checkpoint_id": checkpoint_id,
    }


def score_color(s: int) -> tuple[str, str]:
    """Return (foreground, background) hex color for a score 0-100."""
    if s >= 80:
        return ('#22543d', '#c6f6d5')
    if s >= 55:
        return ('#7b341e', '#feebc8')
    return ('#742a2a', '#fed7d7')


def build_checklist_view(issues: list[dict], deep_audit: bool = False,
                         target_keyword: str = '') -> list[dict]:
    """
    Build the 27-checkpoint status view for a single page.

    For each CHECKLIST item, returns:
      {id, number, label, description, group, status, severity, issue_count, notes}

    Status values:
      'pass'    : no issues raised for this checkpoint
      'fail'    : at least one critical/warning issue
      'info'    : only info-level issues
      'manual'  : checkpoint is auto: 'manual' (always needs human review)
      'skipped' : auto: 'deep' but deep_audit=False; or auto: 'partial' with
                  no signal because optional input (e.g. target_keyword) wasn't given

    The severity field is the worst severity among the issues for that
    checkpoint: 'critical' > 'warning' > 'info'.
    """
    # Group issues by checkpoint_id
    by_cp: dict[str, list[dict]] = defaultdict(list)
    for iss_dict in issues:
        cid = iss_dict.get('checkpoint_id', '')
        if cid:
            by_cp[cid].append(iss_dict)

    sev_rank = {'critical': 3, 'warning': 2, 'info': 1}
    out: list[dict] = []

    for cp in CHECKLIST:
        cid = cp['id']
        related = by_cp.get(cid, [])

        # Determine status
        if cp['auto'] == 'manual':
            status = 'manual'
            severity = 'info'
            notes = 'Cannot be checked automatically — verify manually.'
        elif cp['auto'] == 'deep' and not deep_audit:
            status = 'skipped'
            severity = 'info'
            notes = 'Skipped (deep audit not enabled). Run with "Deep audit" to check.'
        elif cid == 'keyword_usage' and not target_keyword:
            status = 'skipped'
            severity = 'info'
            notes = 'No target keyword provided — fill the "Target keyword" field on Step 1.'
        elif not related:
            status = 'pass'
            severity = 'info'
            notes = ''
        else:
            # Pick worst severity
            worst_rank = max(sev_rank.get(i['severity'], 1) for i in related)
            severity = next(s for s, r in sev_rank.items() if r == worst_rank)
            status = 'info' if severity == 'info' else 'fail'
            # Notes = first issue's title (most relevant)
            notes = related[0]['title']

        out.append({
            'id': cid,
            'number': cp['number'],
            'label': cp['label'],
            'description': cp['description'],
            'group': cp['group'],
            'auto': cp['auto'],
            'status': status,
            'severity': severity,
            'issue_count': len(related),
            'notes': notes,
        })

    return out


def aggregate_checklist(results: list[dict]) -> list[dict]:
    """
    Aggregate per-page checklists into a site-wide view.
    For each checkpoint: count of pages passing / failing / skipped / manual.

    Used by the report's "site-wide checklist" panel.
    """
    counts: dict[str, dict] = {
        cp['id']: {
            'id': cp['id'], 'number': cp['number'], 'label': cp['label'],
            'group': cp['group'], 'description': cp['description'],
            'pass': 0, 'fail': 0, 'info': 0, 'manual': 0, 'skipped': 0,
            'total': 0,
        }
        for cp in CHECKLIST
    }
    for r in results:
        for entry in r.get('checklist', []):
            cid = entry.get('id')
            if cid in counts:
                counts[cid][entry.get('status', 'pass')] += 1
                counts[cid]['total'] += 1

    return [counts[cp['id']] for cp in CHECKLIST]


# ============================================================
#  Executive Summary / Design Suggestions Generator
# ============================================================
#
# This is the "what should you actually DO about it" layer that turns
# raw audit data into a prioritized list of actionable recommendations.
# Used in both the email and a new top section of the HTML report.
#
# Design philosophy: Triage like a senior consultant would.
# - Surface the 3-5 highest-impact issues
# - Group recommendations by theme (SEO, Design, Trust, Content)
# - Quantify scale ("23 pages missing alt text" not just "some pages")
# - Sort by severity × prevalence
# - Phrase as a TODO, not a problem ("Add canonical tags" not "Canonical
#   tag missing")

def _count_issues_by_checkpoint(results: list[dict]) -> dict[str, dict]:
    """Count critical/warning/info issues per checkpoint across all pages."""
    by_cp: dict[str, dict] = defaultdict(
        lambda: {'critical': 0, 'warning': 0, 'info': 0, 'pages': set()}
    )
    for r in results:
        for iss in r.get('issues', []):
            cid = iss.get('checkpoint_id') or 'other'
            sev = iss.get('severity', 'info')
            by_cp[cid][sev] += 1
            by_cp[cid]['pages'].add(r['url'])
    # Convert page sets to counts (sets aren't JSON-serializable)
    return {cid: {**d, 'pages': len(d['pages'])} for cid, d in by_cp.items()}


# Each recommendation has: id, theme, headline, why, action_steps (list),
# affected_pages (count), severity, priority_score (computed).
# The themes are deliberately user-facing: "SEO Foundations" not "seo_seo_seo".

RECOMMENDATION_THEMES = {
    'seo':     'SEO Foundations',
    'design':  'Design & UX',
    'content': 'Content Quality',
    'trust':   'Trust & Conversions',
    'tech':    'Technical Performance',
}

# Map of checkpoint_id -> (theme, design-friendly headline template,
# why-it-matters text, action steps). When we encounter issues for a
# checkpoint we look it up here to produce the recommendation. Anything
# not in this map gets a generic recommendation built from the issue title.

_RECO_TEMPLATES = {
    'page_title': {
        'theme': 'seo',
        'headline': 'Fix missing or weak page titles',
        'why': ("Page titles are the strongest on-page SEO signal and "
                "what users see in search results. Pages without titles, "
                "or with generic titles, rank lower and get clicked less."),
        'actions': [
            'Audit every page title against the 60-70 character sweet spot',
            'Lead with the main keyword, then brand: "SS 304 Pipe | Brand"',
            'Ensure every page has a UNIQUE title (no duplicates across the site)',
        ],
    },
    'meta_description': {
        'theme': 'seo',
        'headline': 'Write compelling meta descriptions',
        'why': ("Meta descriptions don't directly affect rankings, but they "
                "determine click-through rate from search results. Pages "
                "without one get an auto-generated snippet that rarely sells."),
        'actions': [
            'Write a unique 120-158 character description per page',
            'Include a value proposition + light call-to-action',
            'Mention the primary keyword naturally — not stuffed',
        ],
    },
    'h1_tag': {
        'theme': 'seo',
        'headline': 'Standardize one H1 per page',
        'why': ("Multiple H1s or missing H1s confuse Google about what the "
                "page is about. One clear H1 per page is a basic SEO rule."),
        'actions': [
            'Convert extra H1s to H2 (most CMS themes have multiple by mistake)',
            'Ensure every page has at least one H1',
            'Match H1 content to the page title and user intent',
        ],
    },
    'canonical': {
        'theme': 'seo',
        'headline': 'Add canonical tags everywhere',
        'why': ("Without canonicals, Google can't tell which URL is the "
                "'master' version when a page is accessible via multiple "
                "paths. Result: ranking signals get split and you may be "
                "penalized for duplicate content."),
        'actions': [
            'Add <link rel="canonical" href="..."> to every page <head>',
            'Point to the clean, preferred URL (no query params, no trailing slash mismatch)',
            'For paginated pages, canonical points to page 1',
        ],
    },
    'image_alt': {
        'theme': 'design',
        'headline': 'Write descriptive alt text for every image',
        'why': ("Alt text serves three purposes: accessibility (screen readers), "
                "SEO (Google can't 'see' images), and UX (shown when images "
                "fail to load). Missing alt text fails all three."),
        'actions': [
            'Describe what the image SHOWS, not just what it IS',
            'Bad: "image1.jpg". Good: "SS 304 seamless pipe with mill finish"',
            'Decorative images can use empty alt="" (intentional, signals to screen readers to skip)',
        ],
    },
    'image_weight': {
        'theme': 'tech',
        'headline': 'Compress oversized product images',
        'why': ("Images over 100 KB slow page load on mobile networks "
                "(common for B2B buyers in industrial sites). Google uses "
                "page speed as a ranking factor since 2018."),
        'actions': [
            'Use TinyPNG, Squoosh, or ImageOptim to compress without quality loss',
            'Convert to WebP — 30-50% smaller than JPEG at same quality',
            'Add loading="lazy" to images below the fold',
        ],
    },
    'image_filename': {
        'theme': 'seo',
        'headline': 'Rename generic image files to descriptive ones',
        'why': ("Filenames like IMG_1234.jpg or DSC_5678.png are SEO dead "
                "weight. Google reads filenames as ranking signals for image "
                "search and contextually for the page."),
        'actions': [
            'Rename to descriptive, hyphenated lowercase: "ss-304-seamless-pipe.jpg"',
            'Match the filename to the page topic / product',
            'Update any references in code or CMS after renaming',
        ],
    },
    'schema': {
        'theme': 'seo',
        'headline': 'Add Schema.org structured data',
        'why': ("Schema markup is what unlocks rich results in Google: "
                "star ratings, breadcrumbs, FAQ accordions, product pricing. "
                "Pages without schema are invisible to these features."),
        'actions': [
            'Product pages: add Product + Offer schema',
            'FAQ sections: add FAQPage schema',
            'Add BreadcrumbList for navigation hierarchy',
            'Validate at search.google.com/test/rich-results',
        ],
    },
    'breadcrumbs': {
        'theme': 'design',
        'headline': 'Add breadcrumbs with BreadcrumbList schema',
        'why': ("Breadcrumbs help users orient (especially on deep product "
                "pages) and let Google show your site hierarchy in search "
                "results. Free SEO win, immediate UX improvement."),
        'actions': [
            'Add a visible breadcrumb nav: Home > Category > Product',
            'Wrap in <nav aria-label=\"breadcrumb\"> for accessibility',
            'Pair with BreadcrumbList JSON-LD schema',
        ],
    },
    'internal_linking': {
        'theme': 'seo',
        'headline': 'Build a stronger internal linking structure',
        'why': ("Internal links pass PageRank between pages and help Google "
                "discover deep content. Product pages should link to related "
                "products, parent categories, and supporting content."),
        'actions': [
            'On each product page, add a "Related products" section with 3-5 links',
            'Link from blog posts / category pages to specific product pages',
            'Use descriptive anchor text — not "click here"',
        ],
    },
    'cta_button': {
        'theme': 'trust',
        'headline': 'Make CTAs prominent and working',
        'why': ("If buyers can't tell what action to take next — or click a "
                "button that doesn't work — they leave. Every product page "
                "should have a clear 'Get Quote' or 'Contact' CTA above the fold."),
        'actions': [
            'Add a primary CTA in the page header AND another at the bottom',
            'Verify every CTA href points to a real, working URL',
            'Use action-oriented text: "Request Quote" not just "Submit"',
        ],
    },
    'contact_info': {
        'theme': 'trust',
        'headline': 'Display complete contact information',
        'why': ("B2B buyers research thoroughly before reaching out. Missing "
                "phone numbers, missing addresses, or vague contact info "
                "kills trust — buyers go to competitors who look more legitimate."),
        'actions': [
            'Display phone number, email, and physical address in the footer',
            'Add a dedicated /contact page with map and detailed info',
            'For B2B: include direct sales contact names per region',
        ],
    },
    'inquiry_form': {
        'theme': 'trust',
        'headline': 'Add and TEST your inquiry form',
        'why': ("A broken contact form is worse than no form — it silently "
                "loses leads. Most teams discover the form is broken months "
                "after a redesign. Don't be that team."),
        'actions': [
            'Add a contact form with name, email, company, message',
            'Submit a real test inquiry and verify the email arrives',
            'Set up auto-reply so customers know you received their message',
        ],
    },
    'faqs': {
        'theme': 'content',
        'headline': 'Build an FAQ section per product',
        'why': ("FAQ pages capture long-tail search queries that buyers "
                "actually type ('what is the carbon content of SS 304?'). "
                "With FAQPage schema, Google shows your answers directly "
                "in search results — massive visibility boost."),
        'actions': [
            'Ask sales/support for the 5-10 most common pre-purchase questions',
            'Write technically accurate answers (no fluff)',
            'Wrap in FAQPage schema for rich snippets',
        ],
    },
    'mobile_friendly': {
        'theme': 'design',
        'headline': 'Verify the site works on real mobile devices',
        'why': ("Google uses mobile-first indexing — your mobile version IS "
                "your SEO version. If buttons are tiny, text is unreadable, "
                "or tables overflow on mobile, you lose both traffic AND "
                "deals (many B2B buyers research on their phones)."),
        'actions': [
            'Test the site on an actual phone, not just browser DevTools',
            'Tables should scroll horizontally, not overflow',
            'Tap targets should be at least 44px',
            'Don\'t disable zoom — it\'s an accessibility anti-pattern',
        ],
    },
    'specifications': {
        'theme': 'content',
        'headline': 'Add complete specifications tables',
        'why': ("Industrial buyers compare specs side-by-side. Pages without "
                "a specifications table force buyers to email or call for "
                "basic info — most won't bother, they'll just choose a "
                "competitor with the data on the page."),
        'actions': [
            'Every product page: table with size, grade, standard, finish',
            'Use proper HTML <table> markup (not images of tables)',
            'Match column headers to industry standards (NB, OD, schedule, etc.)',
        ],
    },
    'chemical_table': {
        'theme': 'content',
        'headline': 'Verify chemical composition tables match ASTM specs',
        'why': ("Buyers cross-reference your chemical composition table "
                "against ASTM standards. Errors in C/Cr/Ni/Mo percentages "
                "look unprofessional and create legal/QC exposure — buyers "
                "may reject shipments that don't match the published spec."),
        'actions': [
            'Cross-check Carbon, Chromium, Nickel, Molybdenum values vs ASTM A312/A276',
            'Use "max" notation explicitly: "C: 0.08 max"',
            'List values per grade — don\'t mix multiple grades in one row',
        ],
    },
    'mechanical_table': {
        'theme': 'content',
        'headline': 'Add accurate mechanical properties',
        'why': ("Tensile, yield, elongation, and hardness are non-negotiable "
                "data for engineering buyers. Pages without these get "
                "skipped by serious procurement teams."),
        'actions': [
            'Add Tensile Strength, Yield Strength, Elongation, Hardness',
            'Use SI units (MPa) primary, imperial (psi) parenthetical',
            'Cite the ASTM standard the values come from',
        ],
    },
    'equivalent_table': {
        'theme': 'content',
        'headline': 'Add international equivalency tables',
        'why': ("Buyers from Europe, Japan, China, and the Middle East use "
                "different grade nomenclature (EN, DIN, JIS, GB). Pages "
                "without equivalents force international buyers to search "
                "elsewhere — direct revenue loss."),
        'actions': [
            'For each grade, list UNS, EN/DIN, JIS, GB, BS equivalents',
            'Use a reference like outokumpu.com/grades or asminternational.org',
            'Don\'t guess — wrong equivalents are worse than missing ones',
        ],
    },
    'url_structure': {
        'theme': 'seo',
        'headline': 'Clean up URL structures',
        'why': ("Clean URLs (e.g. /products/ss-304-pipe) rank better and "
                "look more trustworthy in search results than database-style "
                "URLs (e.g. ?id=12345). Easier to share and remember too."),
        'actions': [
            'Rewrite database URLs to descriptive slugs',
            'Use hyphens not underscores',
            'Keep URLs short, lowercase, and keyword-rich',
        ],
    },
    'indexability': {
        'theme': 'tech',
        'headline': 'Fix pages set to noindex',
        'why': ("A noindex tag tells Google to NEVER show the page in search "
                "results. Sometimes this is on key product pages by mistake "
                "(left over from staging, or copy-pasted from a template)."),
        'actions': [
            'Search the site for <meta name="robots" content="noindex">',
            'Remove from all pages you WANT in search results',
            'Keep on: thank-you pages, internal admin pages, search results',
        ],
    },
    'subheadings': {
        'theme': 'content',
        'headline': 'Improve content structure with H2/H3 subheadings',
        'why': ("Walls of text don't get read. Subheadings let users scan to "
                "the part they care about, and Google uses them to understand "
                "page structure (often featured snippets come from H2 content)."),
        'actions': [
            'Break content into 4-6 logical sections with H2 headings',
            'Use H3 for sub-sections within each H2',
            'Make subheadings descriptive — they should answer "what\'s in this section?"',
        ],
    },
    'tech_spec_tab': {
        'theme': 'content',
        'headline': 'Provide downloadable datasheets',
        'why': ("Procurement teams routinely save technical datasheets for "
                "internal review. Pages without a Spec Sheet / Datasheet "
                "download lose these high-intent leads."),
        'actions': [
            'Create a PDF datasheet for each product family',
            'Add a "Download Datasheet" button on every product page',
            'Track downloads as conversion events for marketing analytics',
        ],
    },
    'applications': {
        'theme': 'content',
        'headline': 'Make application lists industry-specific',
        'why': ("Generic application lists ('used in many industries') tell "
                "buyers nothing. Specific lists ('petrochemical refineries, "
                "marine heat exchangers, pharma reactors') signal expertise "
                "and help users find products that fit their use case."),
        'actions': [
            'List 5-10 concrete industries or applications per product',
            'Mention specific environments: temperature ranges, corrosive media',
            'Include 1-2 customer use cases as social proof',
        ],
    },
}


def generate_recommendations(results: list[dict]) -> list[dict]:
    """
    Turn audit results into a prioritized list of actionable recommendations.

    Returns a list of recommendation dicts, sorted by priority_score (highest
    first). Priority is severity × prevalence — issues that hit many pages
    AND are critical get surfaced first.

    Each recommendation dict has:
      id              : checkpoint_id (or 'other' for ungrouped)
      theme           : 'seo' | 'design' | 'content' | 'trust' | 'tech'
      theme_label     : Human-readable theme name
      headline        : Imperative action statement
      why             : 1-2 sentences explaining why it matters
      action_steps    : list[str] of concrete steps
      affected_pages  : how many pages of the site have this issue
      critical_count  : critical issues at this checkpoint
      warning_count   : warning issues at this checkpoint
      severity        : highest severity ('critical' | 'warning' | 'info')
      priority_score  : numeric, higher = more important
    """
    if not results:
        return []

    issue_counts = _count_issues_by_checkpoint(results)
    total_pages = len(results)
    recs: list[dict] = []

    for cid, counts in issue_counts.items():
        crit = counts['critical']
        warn = counts['warning']
        info = counts['info']
        pages_affected = counts['pages']

        # Determine severity
        if crit > 0:
            severity = 'critical'
        elif warn > 0:
            severity = 'warning'
        else:
            severity = 'info'

        # Priority score: combines severity and prevalence
        # critical = 10x, warning = 3x, info = 1x; multiplied by % of pages affected
        prevalence = pages_affected / max(1, total_pages)
        sev_weight = 10 if severity == 'critical' else (3 if severity == 'warning' else 1)
        priority_score = sev_weight * (1 + prevalence * 5)

        # Pull template or build generic recommendation
        tpl = _RECO_TEMPLATES.get(cid)
        if tpl:
            theme = tpl['theme']
            headline = tpl['headline']
            why = tpl['why']
            actions = tpl['actions']
        else:
            # Generic fallback: derive from the checkpoint metadata
            cp_meta = CHECKLIST_BY_ID.get(cid)
            if cp_meta:
                theme = 'seo' if cp_meta['group'] == 'On-page SEO' else \
                        'content' if 'Content' in cp_meta['group'] or 'Tables' in cp_meta['group'] else \
                        'design' if cp_meta['group'] in ('Images', 'Page Content') else \
                        'tech'
                headline = f"Address {cp_meta['label'].lower()} issues"
                why = cp_meta['description']
                actions = ['Review the per-page issue list in the full report',
                           f'See checkpoint #{cp_meta["number"]} in the 27-point checklist']
            else:
                continue  # Skip 'other' bucket if no metadata

        recs.append({
            'id': cid,
            'theme': theme,
            'theme_label': RECOMMENDATION_THEMES.get(theme, theme.title()),
            'headline': headline,
            'why': why,
            'action_steps': actions,
            'affected_pages': pages_affected,
            'total_pages': total_pages,
            'critical_count': crit,
            'warning_count': warn,
            'info_count': info,
            'severity': severity,
            'priority_score': priority_score,
        })

    # Sort: highest priority first
    recs.sort(key=lambda r: -r['priority_score'])
    return recs


def top_recommendations(results: list[dict], n: int = 5) -> list[dict]:
    """Convenience: just the top N recommendations for email summaries."""
    return generate_recommendations(results)[:n]


def overall_health_summary(results: list[dict]) -> dict:
    """
    Build a one-glance health summary across the whole site.
    Used at the top of email reports and the executive summary section.
    """
    if not results:
        return {
            'page_count': 0, 'avg_score': 0, 'grade': 'N/A',
            'critical_issues': 0, 'warning_issues': 0, 'info_issues': 0,
            'pages_with_critical': 0,
        }

    page_count = len(results)
    avg_score = round(sum(r['scores'].get('overall', 0) for r in results) / page_count)
    crit_total = sum(sum(1 for i in r['issues'] if i['severity'] == 'critical') for r in results)
    warn_total = sum(sum(1 for i in r['issues'] if i['severity'] == 'warning') for r in results)
    info_total = sum(sum(1 for i in r['issues'] if i['severity'] == 'info') for r in results)
    pages_with_critical = sum(
        1 for r in results
        if any(i['severity'] == 'critical' for i in r['issues'])
    )

    # Letter grade based on average score
    if avg_score >= 90:    grade = 'A'
    elif avg_score >= 80:  grade = 'B'
    elif avg_score >= 70:  grade = 'C'
    elif avg_score >= 60:  grade = 'D'
    else:                  grade = 'F'

    return {
        'page_count': page_count,
        'avg_score': avg_score,
        'grade': grade,
        'critical_issues': crit_total,
        'warning_issues': warn_total,
        'info_issues': info_total,
        'pages_with_critical': pages_with_critical,
    }


# ============================================================
#  Crawler
# ============================================================

class Crawler:
    """
    Same-origin crawler with parallel fetching, link extraction, and progress
    callbacks. Honours a stop signal so the Flask app can cancel mid-crawl.

    Args:
        base_url: starting URL
        max_pages: hard cap on pages to fetch
        threads: parallel fetcher count (1-10)
        delay: pause between batches (seconds)
        timeout: per-request timeout (seconds)
        skip_ssl: bypass SSL verification (for sites with bad certs)
        on_progress: callback(current, total, current_url) - called after each page
        should_stop: callable() -> bool - return True to abort the crawl
    """

    def __init__(
        self,
        base_url: str,
        max_pages: int = 100,
        threads: int = 5,
        delay: float = 0.3,
        timeout: int = 15,
        skip_ssl: bool = False,
        on_progress: Optional[Callable[[int, int, str], None]] = None,
        should_stop: Optional[Callable[[], bool]] = None,
    ):
        parsed = urlparse(base_url)
        self.base_url = base_url.rstrip('/')
        self.base_domain = parsed.netloc
        self.max_pages = max_pages
        self.threads = max(1, min(10, threads))
        self.delay = delay
        self.timeout = timeout
        self.skip_ssl = skip_ssl
        self.on_progress = on_progress or (lambda c, t, u: None)
        self.should_stop = should_stop or (lambda: False)

        self.visited: set[str] = {self.base_url}
        self.q: Queue = Queue()
        self.q.put(self.base_url)
        self.pages: list[dict] = []
        self.lock = threading.Lock()

        self.session = requests.Session()
        self.session.headers['User-Agent'] = (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/121.0 Safari/537.36 audit-wizard/1.0'
        )

    def fetch(self, url: str) -> Optional[dict]:
        """Fetch a single URL and return a page dict (or None for non-HTML)."""
        try:
            t0 = time.time()
            r = self.session.get(
                url, timeout=self.timeout,
                verify=not self.skip_ssl, allow_redirects=True,
            )
            ct = r.headers.get('content-type', '')
            if 'text/html' not in ct.lower():
                return None
            return {
                'url': url,
                'status': r.status_code,
                'html': r.text,
                'response_time': time.time() - t0,
                'content_length': len(r.content),
                # Keep only security-relevant headers (cheap to carry,
                # everything else would bloat memory across many pages)
                'response_headers': {
                    k: v for k, v in r.headers.items()
                    if k.lower() in (
                        'strict-transport-security',
                        'x-content-type-options',
                        'x-frame-options',
                        'content-security-policy',
                        'referrer-policy',
                        'permissions-policy',
                    )
                },
            }
        except Exception as e:
            return {
                'url': url, 'status': 0, 'html': '',
                'response_time': 0, 'error': str(e),
            }

    def extract_links(self, html: str, base: str) -> set[str]:
        """Extract same-domain links from a page, normalised and deduped."""
        soup = BeautifulSoup(html, 'lxml')
        out: set[str] = set()
        for a in soup.find_all('a', href=True):
            href = a['href'].split('#')[0].split('?')[0].strip()
            if not href:
                continue
            full = urljoin(base, href)
            p = urlparse(full)
            if p.netloc != self.base_domain:
                continue
            if any(p.path.lower().endswith(e) for e in SKIP_EXT):
                continue
            if p.scheme not in ('http', 'https'):
                continue
            clean = p.scheme + '://' + p.netloc + p.path.rstrip('/')
            if clean and clean not in self.visited:
                out.add(clean)
        return out

    # --------------------------------------------------------------
    #  Sitemap.xml discovery
    # --------------------------------------------------------------
    # Industrial product sites almost always publish sitemap.xml — they
    # need it for Google indexing of deep product pages. Reading the
    # sitemap gives us much better coverage than menu/footer-only crawling
    # and a known total upfront so the progress bar is accurate.

    def discover_sitemap_urls(self, base_url: str) -> list[str]:
        """
        Discover URLs from sitemap.xml. Returns same-domain URLs only.
        Tries standard sitemap locations + reads robots.txt for custom ones.
        Handles sitemap-index files by recursing one level.
        """
        candidates = [
            urljoin(base_url, '/sitemap.xml'),
            urljoin(base_url, '/sitemap_index.xml'),
            urljoin(base_url, '/sitemap-index.xml'),
        ]
        # Also check robots.txt for Sitemap: directives
        try:
            r = self.session.get(urljoin(base_url, '/robots.txt'),
                                 timeout=10, verify=False)
            if r.status_code == 200:
                for line in r.text.splitlines():
                    if line.lower().startswith('sitemap:'):
                        sm_url = line.split(':', 1)[1].strip()
                        if sm_url and sm_url not in candidates:
                            candidates.append(sm_url)
        except Exception:
            pass

        url_cap = max(self.max_pages * 3, 200)
        all_urls: list[str] = []
        seen: set[str] = set()
        base_domain = urlparse(base_url).netloc

        for sm_url in candidates:
            if len(all_urls) >= url_cap:
                break
            try:
                urls = self._parse_sitemap(
                    sm_url, recurse=True, cap=url_cap - len(all_urls)
                )
                for u in urls:
                    if urlparse(u).netloc == base_domain and u not in seen:
                        seen.add(u)
                        all_urls.append(u)
                if all_urls:
                    break
            except Exception:
                continue

        return all_urls

    def _parse_sitemap(self, sitemap_url: str, recurse: bool = True,
                       cap: int = 1000) -> list[str]:
        """Fetch sitemap.xml and extract <loc> URLs. Recurses sitemap-index."""
        from xml.etree import ElementTree as ET
        try:
            r = self.session.get(sitemap_url, timeout=15, verify=False)
            if r.status_code != 200 or not r.text.strip():
                return []
        except Exception:
            return []
        text = r.text.lstrip('\ufeff').strip()
        try:
            root = ET.fromstring(text)
        except ET.ParseError:
            return []
        ns = {'sm': 'http://www.sitemaps.org/schemas/sitemap/0.9'}
        urls: list[str] = []
        if root.tag.endswith('sitemapindex'):
            if not recurse:
                return []
            for child in root.findall('sm:sitemap/sm:loc', ns):
                if child.text and len(urls) < cap:
                    urls.extend(self._parse_sitemap(
                        child.text.strip(), recurse=False, cap=cap - len(urls)
                    ))
            return urls[:cap]
        for loc in root.findall('sm:url/sm:loc', ns):
            if loc.text and len(urls) < cap:
                urls.append(loc.text.strip())
        return urls

    def _seed_queue_from_sitemap(self):
        """If sitemap.xml exists, use it to seed the URL queue (vs link-following only)."""
        sitemap_urls = self.discover_sitemap_urls(self.base_url)
        if not sitemap_urls:
            return 0  # No sitemap; fall back to menu/footer link discovery
        # Reset queue + visited to use sitemap URLs as the seed
        ordered: list[str] = []
        if self.base_url in sitemap_urls:
            ordered.append(self.base_url)
        for u in sitemap_urls:
            if u != self.base_url and u not in ordered:
                ordered.append(u)
        while not self.q.empty():
            try:
                self.q.get_nowait()
            except Exception:
                break
        self.visited = set(ordered)
        for u in ordered[:self.max_pages]:
            self.q.put(u)
        return len(ordered[:self.max_pages])

    def crawl(self) -> list[dict]:
        """
        Legacy method: collects all pages into a list before returning.
        Kept for backwards compatibility with existing tests.
        For new code, use crawl_streaming() which is much more memory-efficient.
        """
        self._seed_queue_from_sitemap()
        empty_rounds = 0
        while len(self.pages) < self.max_pages and empty_rounds < 4:
            if self.should_stop():
                break

            batch: list[str] = []
            while not self.q.empty() and len(batch) < self.threads:
                try:
                    batch.append(self.q.get_nowait())
                except Exception:
                    break

            if not batch:
                empty_rounds += 1
                time.sleep(0.5)
                continue
            empty_rounds = 0

            with ThreadPoolExecutor(max_workers=len(batch)) as ex:
                futures = {ex.submit(self.fetch, u): u for u in batch}
                for fut in as_completed(futures):
                    if self.should_stop():
                        break
                    res = fut.result()
                    if res and res.get('html') and res.get('status') == 200:
                        with self.lock:
                            if len(self.pages) < self.max_pages:
                                self.pages.append(res)
                                self.on_progress(
                                    len(self.pages), self.max_pages, res['url']
                                )
                                for lk in self.extract_links(res['html'], res['url']):
                                    if lk not in self.visited:
                                        self.visited.add(lk)
                                        self.q.put(lk)

            if self.delay:
                time.sleep(self.delay)

        return self.pages

    def crawl_streaming(self):
        """
        STREAMING crawler — yields one page dict at a time as soon as it's
        fetched. This is the memory-efficient version used in production:
        callers should analyze each page and discard the 'html' field
        immediately, so we never have more than a handful of full HTML
        documents in memory at once.

        Cuts peak memory by ~5× vs the old crawl() pattern that collected
        all pages into a list first. Critical for Render free tier
        (512 MB RAM) when auditing 50+ pages on rich product sites.

        Each yielded dict has the standard shape (url, html, status,
        response_time, content_length). Caller is responsible for
        bookkeeping (counting pages, stopping at max_pages, etc.).
        """
        pages_yielded = 0
        self._seed_queue_from_sitemap()

        empty_rounds = 0
        while pages_yielded < self.max_pages and empty_rounds < 4:
            if self.should_stop():
                return

            batch: list[str] = []
            while not self.q.empty() and len(batch) < self.threads:
                try:
                    batch.append(self.q.get_nowait())
                except Exception:
                    break

            if not batch:
                empty_rounds += 1
                time.sleep(0.5)
                continue
            empty_rounds = 0

            with ThreadPoolExecutor(max_workers=len(batch)) as ex:
                futures = {ex.submit(self.fetch, u): u for u in batch}
                for fut in as_completed(futures):
                    if self.should_stop():
                        return
                    res = fut.result()
                    if res and res.get('html') and res.get('status') == 200:
                        with self.lock:
                            if pages_yielded >= self.max_pages:
                                return
                            pages_yielded += 1
                            self.on_progress(
                                pages_yielded, self.max_pages, res['url']
                            )
                            # Extract links BEFORE yielding (so we still queue
                            # discovered links even though caller will discard
                            # the HTML after analyzing).
                            for lk in self.extract_links(res['html'], res['url']):
                                if lk not in self.visited:
                                    self.visited.add(lk)
                                    self.q.put(lk)
                        # Yield outside the lock — caller does heavy work
                        yield res

            if self.delay:
                time.sleep(self.delay)


# ============================================================
#  Page analyzer
# ============================================================

class Analyzer:
    """
    Analyse a fetched page for SEO, HTML, performance, content, and
    industry-specific issues, plus the 27-point checklist.

    Options:
      industry         : 'auto' | 'metals' | 'ecommerce' | 'saas' | 'healthcare'
                         | 'realestate' | 'generic'
      target_keyword   : Optional. If provided, runs keyword-density and
                         keyword-placement checks (checklist item 13).
      deep_audit       : If True, runs slow checks that require extra
                         network requests:
                           - Image size HEAD requests (item 11)
                           - CTA link validation (item 22)
                         Default False because these can add minutes to a
                         100-page audit.
    """

    def __init__(self, industry: str = 'auto',
                 target_keyword: str = '',
                 deep_audit: bool = False):
        self.industry = industry
        self.target_keyword = (target_keyword or '').strip()
        self.deep_audit = bool(deep_audit)

    def analyze(self, pg: dict) -> dict:
        url = pg.get('url', '')
        html = pg.get('html', '')
        if not html:
            return {
                'url': url, 'title': '[blocked]', 'issues': [], 'scores': {},
                'word_count': 0, 'grades_found': [], 'response_time': 0,
                'industry': '—', 'status': pg.get('status', 0),
            }

        soup = BeautifulSoup(html, 'lxml')
        text = soup.get_text(' ', strip=True)
        ind = self.industry if self.industry != 'auto' else detect_industry(url, text)

        issues: list[dict] = []
        scores: dict[str, int] = {}

        # Universal checks (every industry gets these)
        r = self._check_seo(soup, text, url)
        issues += r['i']; scores['seo'] = r['s']
        r = self._check_html(soup, url)
        issues += r['i']; scores['html'] = r['s']
        r = self._check_performance(pg, soup)
        issues += r['i']; scores['performance'] = r['s']
        r = self._check_content(soup, text, ind)
        issues += r['i']; scores['content'] = r['s']

        # NEW: Checklist-driven checks (run for every page)
        issues += self._check_url_structure(url)
        issues += self._check_breadcrumbs(soup)
        issues += self._check_image_filenames(soup)
        issues += self._check_internal_linking(soup, url)
        issues += self._check_keyword_usage(soup, text, self.target_keyword)
        issues += self._check_product_tables(soup, text)
        issues += self._check_applications(soup, text, ind)
        issues += self._check_cta_buttons(soup, url, deep=self.deep_audit)
        issues += self._check_faqs(soup, text)
        issues += self._check_schema_validation(soup)
        issues += self._check_responsive(soup, html)
        issues += self._check_inquiry_form(soup)
        # 360-degree checks: security headers, mixed content, social tags
        issues += self._check_security_headers(pg)
        issues += self._check_mixed_content(soup, url)
        issues += self._check_social_media_tags(soup)

        # Deep-audit-only network checks
        if self.deep_audit:
            issues += self._check_image_weights(soup, url)

        # Industry-specific checks
        if ind == 'metals':
            grades = list(set(GRADE_RE.findall(text)))
            issues += self._check_chemical(text, grades)
            issues += self._check_mechanical(text, grades)
            issues += self._check_astm_standards(text, url)
            issues += self._check_equivalents(text, grades)
            scores['chemical']   = max(0, 100 - len([i for i in issues if i['category'] == 'chemical']) * 15)
            scores['mechanical'] = max(0, 100 - len([i for i in issues if i['category'] == 'mechanical']) * 15)
            scores['standards']  = max(0, 100 - len([i for i in issues if i['category'] == 'astm']) * 12)
            scores['equivalent'] = max(0, 100 - len([i for i in issues if i['category'] == 'equivalent']) * 10)
        elif ind == 'ecommerce':
            r = self._check_ecommerce(soup, text)
            issues += r['i']; scores['ecommerce'] = r['s']
        elif ind == 'saas':
            r = self._check_saas(soup, text)
            issues += r['i']; scores['saas'] = r['s']
        elif ind == 'healthcare':
            r = self._check_healthcare(soup, text)
            issues += r['i']; scores['healthcare'] = r['s']
        elif ind == 'realestate':
            r = self._check_realestate(soup, text)
            issues += r['i']; scores['realestate'] = r['s']

        scores['overall'] = round(sum(scores.values()) / max(1, len(scores)))

        title_tag = soup.find('title')
        title = title_tag.get_text(strip=True) if title_tag else ''
        grades_found = list(set(GRADE_RE.findall(text))) if ind == 'metals' else []

        # Build the per-page checklist view: status of all 27 checkpoints
        checklist_view = build_checklist_view(issues, deep_audit=self.deep_audit,
                                              target_keyword=self.target_keyword)

        return {
            'url': url, 'title': title, 'issues': issues, 'scores': scores,
            'word_count': len(text.split()), 'grades_found': grades_found,
            'response_time': pg.get('response_time', 0),
            'industry': ind, 'status': pg.get('status', 200),
            'checklist': checklist_view,
        }

    # ---------- SEO ----------
    def _check_seo(self, soup, text, url) -> dict:
        issues: list[dict] = []
        score = 100

        title_tag = soup.find('title')
        title = title_tag.get_text(strip=True) if title_tag else ''
        if not title:
            issues.append(_iss('seo', 'critical', 'Title tag missing',
                               'Most basic SEO element. Search results show this.',
                               '<title>Product Name | Brand</title>',
                               checkpoint_id='page_title'))
            score -= 25
        elif len(title) < 30:
            issues.append(_iss('seo', 'warning', f'Title too short ({len(title)} chars)',
                               '50-60 characters is ideal for search snippets.',
                               f'Expand: "{title} | Grade | Company"',
                               checkpoint_id='page_title'))
            score -= 10
        elif len(title) > 65:
            issues.append(_iss('seo', 'warning', f'Title too long ({len(title)} chars)',
                               'Google truncates after ~65 characters.',
                               'Trim to 55-60 chars',
                               checkpoint_id='page_title'))
            score -= 5

        md = soup.find('meta', attrs={'name': re.compile(r'^description$', re.I)})
        desc = md.get('content', '').strip() if md else ''
        if not desc:
            issues.append(_iss('seo', 'critical', 'Meta description missing',
                               'No search snippet will be shown.',
                               '<meta name="description" content="150 char description...">',
                               checkpoint_id='meta_description'))
            score -= 20
        elif len(desc) < 80:
            issues.append(_iss('seo', 'warning', f'Meta description too short ({len(desc)} chars)',
                               '120-158 chars is ideal.',
                               f'Expand: "{desc[:50]}..."',
                               checkpoint_id='meta_description'))
            score -= 8
        elif len(desc) > 160:
            issues.append(_iss('seo', 'info', f'Meta description too long ({len(desc)} chars)',
                               'Google truncates after ~158 chars.', 'Trim it down',
                               checkpoint_id='meta_description'))
            score -= 3

        h1s = soup.find_all('h1')
        if not h1s:
            issues.append(_iss('seo', 'critical', 'H1 missing',
                               'Every page needs one main heading.',
                               '<h1>Main Page Heading</h1>',
                               checkpoint_id='h1_tag'))
            score -= 20
        elif len(h1s) > 1:
            issues.append(_iss('seo', 'warning', f'{len(h1s)} H1 tags (only 1 needed)',
                               'Multiple H1s confuse search engines.',
                               'Convert extra H1s to H2 or H3',
                               checkpoint_id='h1_tag'))
            score -= 10

        if not soup.find('link', rel='canonical'):
            issues.append(_iss('seo', 'warning', 'Canonical link missing',
                               'Risk of duplicate content penalties.',
                               f'<link rel="canonical" href="{url}">',
                               checkpoint_id='canonical'))
            score -= 8

        if not soup.find('meta', property='og:title'):
            issues.append(_iss('seo', 'info', 'Open Graph tags missing',
                               'No social media share preview.',
                               '<meta property="og:title" content="...">\n'
                               '<meta property="og:image" content="...">',
                               checkpoint_id='schema'))
            score -= 5

        if not soup.find('script', type='application/ld+json'):
            issues.append(_iss('seo', 'warning', 'Schema markup missing',
                               'No rich snippets in search results.',
                               '{"@context":"https://schema.org","@type":"WebPage","name":"..."}',
                               checkpoint_id='schema'))
            score -= 10

        no_alt = [i for i in soup.find_all('img') if not i.get('alt', '').strip()]
        if no_alt:
            issues.append(_iss('seo', 'warning',
                               f'{len(no_alt)} images without alt text',
                               'Image SEO and accessibility are reduced.',
                               '<img src="..." alt="Descriptive text">',
                               checkpoint_id='image_alt'))
            score -= min(15, len(no_alt) * 2)

        rob = soup.find('meta', attrs={'name': re.compile(r'^robots$', re.I)})
        if rob and 'noindex' in (rob.get('content', '') or '').lower():
            issues.append(_iss('seo', 'critical', 'Page is set to NOINDEX',
                               'Google will NOT index this page!',
                               'Remove noindex or change to "index,follow"',
                               checkpoint_id='indexability'))
            score -= 40

        if not soup.find_all('h2'):
            issues.append(_iss('seo', 'info', 'No H2 headings',
                               'Content structure is weak.',
                               '<h2>Section Heading</h2>',
                               checkpoint_id='subheadings'))
            score -= 5

        return {'i': issues, 's': max(0, score)}

    # ---------- HTML ----------
    def _check_html(self, soup, url) -> dict:
        issues: list[dict] = []
        score = 100

        if not soup.find('meta', charset=True):
            issues.append(_iss('html', 'warning', 'Charset missing',
                               'Encoding issues may occur.',
                               '<meta charset="UTF-8">',
                               checkpoint_id='mobile_friendly'))
            score -= 8

        if not soup.find('meta', attrs={'name': 'viewport'}):
            issues.append(_iss('html', 'critical', 'Viewport meta missing',
                               'Mobile rendering will be broken.',
                               '<meta name="viewport" content="width=device-width, initial-scale=1.0">',
                               checkpoint_id='mobile_friendly'))
            score -= 20

        if not url.startswith('https'):
            issues.append(_iss('html', 'critical', 'No HTTPS',
                               'No SSL — Google penalises non-HTTPS sites.',
                               "Enable Let's Encrypt SSL on your hosting (free)"))
            score -= 25

        if soup.find(string=re.compile(r'lorem ipsum', re.I)):
            issues.append(_iss('html', 'warning', 'Lorem ipsum text found',
                               'Draft placeholder content is live.',
                               'Replace with real content'))
            score -= 15

        for f in soup.find_all('form'):
            if not f.get('action', ''):
                issues.append(_iss('html', 'warning', 'Form action missing',
                                   'Form will not submit anywhere.',
                                   '<form action="/submit" method="POST">',
                                   checkpoint_id='inquiry_form'))
                score -= 10
                break

        if not (soup.find('link', rel=re.compile('icon', re.I))
                or soup.find('link', rel='shortcut icon')):
            issues.append(_iss('html', 'info', 'Favicon missing',
                               'No tab icon will appear in browsers.',
                               '<link rel="icon" href="/favicon.ico">'))
            score -= 5

        return {'i': issues, 's': max(0, score)}

    # ---------- Performance ----------
    def _check_performance(self, pg, soup) -> dict:
        issues: list[dict] = []
        score = 100
        rt = pg.get('response_time', 0)
        sz = pg.get('content_length', 0)

        if rt > 3:
            issues.append(_iss('performance', 'critical',
                               f'Slow page load ({rt:.1f}s)',
                               'Google ranks slower than 3s lower.',
                               'Compress images, use CDN, enable caching'))
            score -= 30
        elif rt > 1.5:
            issues.append(_iss('performance', 'warning',
                               f'Page load somewhat slow ({rt:.1f}s)',
                               '1-3s is improvable.',
                               'Lazy-load images, compress assets'))
            score -= 15

        if sz > 500000:
            issues.append(_iss('performance', 'warning',
                               f'Page size large ({sz // 1024} KB)',
                               '500+ KB is slow on mobile networks.',
                               'Compress images, remove unused CSS/JS'))
            score -= 10

        return {'i': issues, 's': max(0, score)}

    # ---------- Content quality ----------
    def _check_content(self, soup, text, ind) -> dict:
        issues: list[dict] = []
        score = 100
        words = len(text.split())

        if words < 300:
            issues.append(_iss('content', 'warning',
                               f'Thin content ({words} words)',
                               'Aim for 500+ words on key pages.',
                               'Add specifications, FAQs, applications'))
            score -= 20

        profile = INDUSTRY_PROFILES.get(ind, INDUSTRY_PROFILES['generic'])
        for sec in profile['required_sections']:
            if sec not in text.lower():
                issues.append(_iss('content', 'info',
                                   f'"{sec.title()}" section missing',
                                   f'Important section for {ind} sites.',
                                   f'Add a "{sec.title()}" section'))
                score -= 5

        has_contact = bool(re.search(
            r'(\+?[\d\s\-\(\)]{10,15})|([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+)', text
        ))
        if not has_contact:
            issues.append(_iss('content', 'warning',
                               'Contact info missing (no phone/email visible)',
                               'Buyers cannot contact you.',
                               'Display phone and email clearly',
                               checkpoint_id='contact_info'))
            score -= 15

        ctas = ['contact', 'enquire', 'quote', 'buy now', 'order', 'call us',
                'whatsapp', 'request', 'get started', 'free trial', 'book', 'subscribe']
        if not any(c in text.lower() for c in ctas):
            issues.append(_iss('content', 'warning', 'No CTA detected',
                               'Users have no clear next step.',
                               'Add: "Get Quote" / "Contact Us" / "Enquire Now"',
                               checkpoint_id='cta_button'))
            score -= 10

        return {'i': issues, 's': max(0, score)}

    # ============================================================
    #  NEW CHECKLIST-DRIVEN CHECKS
    #  Each method covers one or more of the 27 checkpoints.
    # ============================================================

    # ---------- Checklist #1: URL Structure ----------
    def _check_url_structure(self, url) -> list[dict]:
        """
        Heuristics for "clean, readable, keyword-rich URL":
          - No URL params like ?id=12345
          - Reasonable length (< 100 chars)
          - Lowercase
          - Hyphens, not underscores or spaces
          - Not too deep (< 5 path segments)
        """
        issues: list[dict] = []
        parsed = urlparse(url)
        path = parsed.path or '/'

        if parsed.query and re.search(r'(\?|&)(id|p|q|page)=\d+', url, re.I):
            issues.append(_iss('seo', 'warning', 'URL uses query parameters',
                               'Database-style URLs (?id=123) are bad for SEO.',
                               'Rewrite to clean URLs like /products/ss-304-pipe',
                               checkpoint_id='url_structure'))

        if len(url) > 100:
            issues.append(_iss('seo', 'info', f'URL very long ({len(url)} chars)',
                               'Long URLs are harder to share and rank slightly worse.',
                               'Aim for under 80 characters where possible',
                               checkpoint_id='url_structure'))

        if '_' in path:
            issues.append(_iss('seo', 'info', 'Underscores in URL',
                               'Google treats hyphens as word separators, not underscores.',
                               'Use ss-304-pipe.html, not ss_304_pipe.html',
                               checkpoint_id='url_structure'))

        if any(c.isupper() for c in path):
            issues.append(_iss('seo', 'info', 'Uppercase letters in URL',
                               'Mixed-case URLs cause duplicate-content risks.',
                               'Use lowercase URLs only',
                               checkpoint_id='url_structure'))

        if path.count('/') > 5:
            issues.append(_iss('seo', 'info', f'URL deeply nested ({path.count("/")} levels)',
                               'Deeply-nested URLs rank worse and confuse users.',
                               'Flatten to 2-3 levels: /products/ss-304-pipe',
                               checkpoint_id='url_structure'))

        return issues

    # ---------- Checklist #7: Breadcrumbs ----------
    def _check_breadcrumbs(self, soup) -> list[dict]:
        """
        Detect breadcrumbs by either:
          - BreadcrumbList JSON-LD schema
          - Visible breadcrumb HTML (nav with class containing "breadcrumb",
            or ol/ul with "breadcrumb" in class)
        """
        issues: list[dict] = []

        # JSON-LD BreadcrumbList
        has_schema = False
        for s in soup.find_all('script', type='application/ld+json'):
            content = s.string or ''
            if 'BreadcrumbList' in content:
                has_schema = True
                break

        # Visible breadcrumb HTML
        has_visible = bool(
            soup.find('nav', attrs={'class': re.compile(r'breadcrumb', re.I)}) or
            soup.find(['ol', 'ul'], attrs={'class': re.compile(r'breadcrumb', re.I)}) or
            soup.find(attrs={'aria-label': re.compile(r'breadcrumb', re.I)})
        )

        if not has_visible and not has_schema:
            issues.append(_iss('seo', 'warning', 'Breadcrumbs missing',
                               'Breadcrumbs help users and Google understand site structure.',
                               'Add <nav aria-label="breadcrumb"> with links + BreadcrumbList JSON-LD',
                               checkpoint_id='breadcrumbs'))
        elif not has_schema:
            issues.append(_iss('seo', 'info', 'Breadcrumb schema missing',
                               'Visible breadcrumbs found but no BreadcrumbList schema.',
                               'Add BreadcrumbList JSON-LD so Google shows them in search results',
                               checkpoint_id='breadcrumbs'))
        return issues

    # ---------- Checklist #9: Image File Names ----------
    def _check_image_filenames(self, soup) -> list[dict]:
        """
        Flag generic / camera-default image filenames:
          - IMG_1234, DSC_1234, photo123, image1, 1.jpg, etc.
          - All-numeric or random-string filenames
        """
        issues: list[dict] = []
        bad_patterns = re.compile(
            r'^(img[_\-]?\d+|dsc[_\-]?\d+|photo\d*|image\d+|pic\d+|'
            r'\d+|untitled|screenshot|temp|new|copy|file)$', re.I
        )
        bad_count = 0
        examples: list[str] = []

        for img in soup.find_all('img'):
            src = (img.get('src') or '').strip()
            if not src or src.startswith('data:'):
                continue
            # Extract just the filename without extension
            fname = src.rsplit('/', 1)[-1].rsplit('.', 1)[0]
            if bad_patterns.match(fname):
                bad_count += 1
                if len(examples) < 3:
                    examples.append(fname)

        if bad_count:
            issues.append(_iss('seo', 'warning',
                               f'{bad_count} images have generic filenames',
                               f'Filenames like "{", ".join(examples)}" miss SEO opportunity.',
                               'Rename to descriptive filenames: ss-304-seamless-pipe.jpg',
                               checkpoint_id='image_filename'))
        return issues

    # ---------- Checklist #11: Image Optimization (deep mode only) ----------
    def _check_image_weights(self, soup, base_url) -> list[dict]:
        """
        For each image on the page, send a HEAD request to check size.
        SLOW: adds ~1 request per image. Only runs when deep_audit=True.

        Limits: max 30 images per page, 5s timeout each, 3s parallel pool.
        """
        issues: list[dict] = []
        srcs: list[str] = []
        for img in soup.find_all('img'):
            src = (img.get('src') or '').strip()
            if not src or src.startswith('data:'):
                continue
            full = urljoin(base_url, src)
            if full not in srcs:
                srcs.append(full)
            if len(srcs) >= 30:
                break  # cap at 30 images

        if not srcs:
            return issues

        oversized: list[tuple[str, int]] = []
        try:
            with ThreadPoolExecutor(max_workers=5) as ex:
                futs = {ex.submit(self._image_size, u): u for u in srcs}
                for f in as_completed(futs, timeout=20):
                    try:
                        u = futs[f]
                        size = f.result()
                        if size and size > 100 * 1024:  # > 100 KB
                            oversized.append((u, size))
                    except Exception:
                        continue
        except Exception:
            pass  # don't fail the audit if HEAD requests are blocked

        if oversized:
            top3 = sorted(oversized, key=lambda x: -x[1])[:3]
            example = ', '.join(f'{u.split("/")[-1]} ({s // 1024}KB)' for u, s in top3)
            issues.append(_iss('performance', 'warning',
                               f'{len(oversized)} images over 100 KB',
                               f'Heavy images slow page load. Largest: {example}',
                               'Compress with TinyPNG / Squoosh, convert to WebP',
                               checkpoint_id='image_weight'))
        return issues

    def _image_size(self, url: str) -> Optional[int]:
        """HEAD request returning content length, or None on failure."""
        try:
            r = requests.head(url, timeout=5, allow_redirects=True, verify=False)
            cl = r.headers.get('content-length')
            return int(cl) if cl else None
        except Exception:
            return None

    # ---------- Checklist #12: Internal Linking ----------
    def _check_internal_linking(self, soup, base_url) -> list[dict]:
        """
        Count internal links (same-domain anchors with content), excluding
        navigation-only links (very short text, # anchors, mailto:, tel:).
        """
        issues: list[dict] = []
        try:
            base_domain = urlparse(base_url).netloc
        except Exception:
            return issues

        substantive_links = 0
        for a in soup.find_all('a', href=True):
            href = a['href'].strip()
            if not href or href.startswith(('#', 'mailto:', 'tel:', 'javascript:')):
                continue
            text = a.get_text(strip=True)
            if len(text) < 4:  # skip "Home", "→", etc — too short to be content links
                continue
            try:
                full = urljoin(base_url, href)
                if urlparse(full).netloc == base_domain:
                    substantive_links += 1
            except Exception:
                continue

        if substantive_links < 3:
            issues.append(_iss('seo', 'warning',
                               f'Few internal links ({substantive_links})',
                               'Pages should link to related products / categories for SEO.',
                               'Add 5-10 contextual links to related pages',
                               checkpoint_id='internal_linking'))
        return issues

    # ---------- Checklist #13: Keyword Usage ----------
    def _check_keyword_usage(self, soup, text, target_keyword: str) -> list[dict]:
        """
        If user provided a target keyword:
         - Check it's in the page title
         - Check it's in the H1
         - Check density 0.5%-3% (warn if zero, warn if >3% = stuffed)
         - Check it's in URL slug
        If no keyword provided, this method is skipped by analyze().
        """
        issues: list[dict] = []
        if not target_keyword or not text:
            return issues

        kw = target_keyword.lower().strip()
        text_lower = text.lower()

        # Density
        words = re.findall(r'\b\w+\b', text_lower)
        kw_count = len(re.findall(r'\b' + re.escape(kw) + r'\b', text_lower))
        density = (kw_count / max(1, len(words))) * 100

        if kw_count == 0:
            issues.append(_iss('content', 'critical',
                               f'Target keyword "{target_keyword}" not in page content',
                               'The page does not mention the keyword you are targeting.',
                               f'Naturally include "{target_keyword}" 3-5 times in the body',
                               checkpoint_id='keyword_usage'))
        elif density > 3:
            issues.append(_iss('content', 'warning',
                               f'Keyword over-used ({density:.1f}% density)',
                               'Google penalises keyword stuffing.',
                               'Reduce to 1-3% density (around 1 mention per 50-100 words)',
                               checkpoint_id='keyword_usage'))
        elif density < 0.3:
            issues.append(_iss('content', 'info',
                               f'Keyword sparse ({density:.2f}% density, {kw_count} mentions)',
                               'Slightly more keyword usage may help.',
                               f'Aim for 0.5-2% density',
                               checkpoint_id='keyword_usage'))

        # Title
        title_tag = soup.find('title')
        title = title_tag.get_text(strip=True).lower() if title_tag else ''
        if title and kw not in title:
            issues.append(_iss('seo', 'warning',
                               f'Keyword "{target_keyword}" not in page title',
                               'Title is the strongest on-page SEO signal.',
                               f'Add "{target_keyword}" to the <title> tag',
                               checkpoint_id='keyword_usage'))

        # H1
        h1 = soup.find('h1')
        h1_text = h1.get_text(strip=True).lower() if h1 else ''
        if h1_text and kw not in h1_text:
            issues.append(_iss('seo', 'info',
                               f'Keyword "{target_keyword}" not in H1',
                               'H1 should reinforce the main keyword.',
                               f'Include "{target_keyword}" in the H1 heading',
                               checkpoint_id='keyword_usage'))

        return issues

    # ---------- Checklist #15, #19, #20: Specifications / Size / Tech-spec tables ----------
    def _check_product_tables(self, soup, text) -> list[dict]:
        """
        Detect presence of:
          - Specifications table (sizes, grades, standards)
          - Size / dimension table
          - Technical specification tab/section
        We look at table headers and section headings.
        """
        issues: list[dict] = []
        text_lower = text.lower()

        # Pull all table header text
        table_headers: list[str] = []
        for table in soup.find_all('table'):
            for th in table.find_all(['th', 'td'])[:8]:  # first row usually has headers
                table_headers.append(th.get_text(strip=True).lower())

        all_headers = ' '.join(table_headers)

        # #15: Specifications table — should have sizes/grades/standards keywords
        spec_signals = ['size', 'grade', 'standard', 'specification', 'spec',
                        'thickness', 'diameter', 'schedule']
        if not soup.find_all('table'):
            issues.append(_iss('content', 'warning', 'No tables on page',
                               'Product pages need at least one specification table.',
                               'Add a table with sizes, grades, and standards',
                               checkpoint_id='specifications'))
        elif not any(s in all_headers for s in spec_signals):
            issues.append(_iss('content', 'info',
                               'Tables present but no specifications table detected',
                               'Found tables but none with size/grade/standard headers.',
                               'Add a specifications table with size, grade, standard columns',
                               checkpoint_id='specifications'))

        # #19: Size/dimension table — separate check
        size_signals = ['size', 'dimension', 'od', 'id ', 'inch', 'mm', 'nb',
                        'wall thickness', 'diameter']
        has_size_table = (
            'dimension' in text_lower or 'size chart' in text_lower or
            any(s in all_headers for s in ['od', 'id ', 'nb', 'mm', 'inch'])
        )
        if soup.find_all('table') and not has_size_table:
            issues.append(_iss('content', 'info', 'Size / dimension table missing',
                               'Industrial product pages should show exact dimensions.',
                               'Add a size table: NB | OD | Wall Thickness | Schedule',
                               checkpoint_id='size_table'))

        # #20: Technical specification tab — look for Spec Sheet / Datasheet / Tech Spec
        tech_signals = ['datasheet', 'data sheet', 'technical specification',
                        'spec sheet', 'specification sheet']
        if not any(s in text_lower for s in tech_signals):
            issues.append(_iss('content', 'info',
                               'No technical specification / datasheet reference',
                               'Buyers expect a downloadable datasheet or detailed tech spec section.',
                               'Add a "Datasheet" link or "Technical Specification" tab',
                               checkpoint_id='tech_spec_tab'))

        return issues

    # ---------- Checklist #21: Applications ----------
    def _check_applications(self, soup, text, industry) -> list[dict]:
        """
        Detect an Applications section. We can check:
          - Heading containing "applications" / "industries" / "uses"
          - List items in that section
        We can't judge "industry-specific" without LLM — so we just verify
        a section exists and isn't a tiny placeholder.
        """
        issues: list[dict] = []
        text_lower = text.lower()

        # Look for an Applications-style heading
        app_heading = None
        for h in soup.find_all(['h2', 'h3', 'h4']):
            ht = h.get_text(strip=True).lower()
            if any(kw in ht for kw in ['application', 'industries', 'usage', 'uses', 'where used']):
                app_heading = h
                break

        if not app_heading and 'application' not in text_lower:
            issues.append(_iss('content', 'warning', 'No Applications section',
                               'Buyers want to know what the product is used for.',
                               'Add an "Applications" section listing industries served',
                               checkpoint_id='applications'))
            return issues

        # If heading found, check the following list has at least 3 items
        if app_heading:
            # Look for the next ul/ol after this heading
            next_list = app_heading.find_next(['ul', 'ol'])
            if next_list:
                items = next_list.find_all('li')
                if len(items) < 3:
                    issues.append(_iss('content', 'info',
                                       f'Applications list very short ({len(items)} items)',
                                       'Aim for 5-10 specific applications.',
                                       'List specific industries / use cases',
                                       checkpoint_id='applications'))

        return issues

    # ---------- Checklist #22: CTA Button (deep mode validates link) ----------
    def _check_cta_buttons(self, soup, base_url, deep: bool) -> list[dict]:
        """
        Find CTA buttons (anchor tags with button-ish classes or CTA text).
        In deep mode, send HEAD requests to validate links work.
        """
        issues: list[dict] = []
        cta_keywords = re.compile(
            r'\b(get\s+quote|enquire|enquiry|contact\s+us|request|book|buy|order|'
            r'download|free\s+trial|get\s+started|sign\s+up|subscribe|whatsapp|call\s+now)\b',
            re.I
        )

        ctas: list = []
        # Anchors with button class or CTA text
        for a in soup.find_all('a', href=True):
            text = a.get_text(strip=True)
            classes = ' '.join(a.get('class', [])).lower()
            if 'btn' in classes or 'button' in classes or cta_keywords.search(text):
                ctas.append(a)
                if len(ctas) >= 5:
                    break

        if not ctas:
            issues.append(_iss('content', 'warning', 'No CTA button found',
                               'Page has no clear call-to-action.',
                               'Add a primary CTA: "Get Quote" / "Enquire Now" / "Contact Us"',
                               checkpoint_id='cta_button'))
            return issues

        # Validate hrefs (always)
        for cta in ctas:
            href = (cta.get('href') or '').strip()
            text = cta.get_text(strip=True)[:30]

            if href in ('#', '', 'javascript:void(0)'):
                issues.append(_iss('content', 'critical',
                                   f'CTA "{text}" has empty/dead link',
                                   'CTA button does not go anywhere.',
                                   f'Set href to a real URL like /contact or /quote',
                                   checkpoint_id='cta_button'))

        # Deep mode: HEAD-check each CTA's target URL
        if deep:
            for cta in ctas:
                href = (cta.get('href') or '').strip()
                if not href or href.startswith(('#', 'javascript:', 'mailto:', 'tel:')):
                    continue
                full = urljoin(base_url, href)
                try:
                    r = requests.head(full, timeout=5, allow_redirects=True, verify=False)
                    if r.status_code >= 400:
                        text = cta.get_text(strip=True)[:30]
                        issues.append(_iss('content', 'critical',
                                           f'CTA "{text}" link returns {r.status_code}',
                                           f'Target URL is broken: {full}',
                                           'Fix or update the link',
                                           checkpoint_id='cta_button'))
                except Exception:
                    pass  # don't fail audit if HEAD blocks

        return issues

    # ---------- Checklist #23: FAQs ----------
    def _check_faqs(self, soup, text) -> list[dict]:
        """
        Detect FAQs by:
          - FAQPage / Question schema
          - Heading text containing "FAQ" / "Frequently Asked"
          - Accordion patterns (details/summary tags)
        """
        issues: list[dict] = []

        # Schema check
        has_faq_schema = False
        for s in soup.find_all('script', type='application/ld+json'):
            content = s.string or ''
            if 'FAQPage' in content or '"@type":"Question"' in content.replace(' ', ''):
                has_faq_schema = True
                break

        # Heading check
        faq_heading = False
        for h in soup.find_all(['h2', 'h3', 'h4']):
            if re.search(r'\bfaq|frequently\s+asked\b', h.get_text(strip=True), re.I):
                faq_heading = True
                break

        # Accordion check (semantic HTML5)
        has_accordion = bool(soup.find('details'))

        if not (has_faq_schema or faq_heading or has_accordion):
            issues.append(_iss('content', 'info', 'No FAQ section',
                               'FAQ pages capture long-tail search queries and rank well.',
                               'Add 5-10 common product questions with FAQPage schema',
                               checkpoint_id='faqs'))
        elif (faq_heading or has_accordion) and not has_faq_schema:
            issues.append(_iss('seo', 'info', 'FAQ section without FAQPage schema',
                               'Visible FAQ section but no schema — losing rich-result eligibility.',
                               'Wrap FAQ in FAQPage JSON-LD for Google rich snippets',
                               checkpoint_id='faqs'))
        return issues

    # ---------- Checklist #25: Schema validation ----------
    def _check_schema_validation(self, soup) -> list[dict]:
        """
        Beyond presence, validate that JSON-LD parses cleanly and contains
        sensible @type values. Doesn't check schema correctness exhaustively
        (that requires schema.org's own validator) — just basic sanity.
        """
        issues: list[dict] = []
        import json

        scripts = soup.find_all('script', type='application/ld+json')
        if not scripts:
            return issues  # already flagged by _check_seo

        broken = 0
        types_found: list[str] = []
        for s in scripts:
            content = s.string or ''
            if not content.strip():
                broken += 1
                continue
            try:
                data = json.loads(content)
                # Could be a single object or a list
                items = data if isinstance(data, list) else [data]
                for item in items:
                    if isinstance(item, dict):
                        t = item.get('@type', '')
                        if t:
                            types_found.append(t if isinstance(t, str) else str(t))
            except (json.JSONDecodeError, TypeError):
                broken += 1

        if broken:
            issues.append(_iss('seo', 'warning',
                               f'{broken} JSON-LD blocks have parse errors',
                               'Broken schema is invisible to search engines.',
                               'Validate at search.google.com/test/rich-results',
                               checkpoint_id='schema'))

        return issues

    # ---------- Checklist #26: Mobile / Responsive ----------
    def _check_responsive(self, soup, html) -> list[dict]:
        """
        Beyond viewport meta (already checked in _check_html), look for
        responsive CSS hints in inline/linked styles.
        """
        issues: list[dict] = []

        # Check viewport content is sensible (not just present)
        vp = soup.find('meta', attrs={'name': 'viewport'})
        if vp:
            content = vp.get('content', '')
            if 'width=device-width' not in content:
                issues.append(_iss('html', 'warning',
                                   'Viewport meta lacks width=device-width',
                                   'Mobile sizing will be broken.',
                                   '<meta name="viewport" content="width=device-width, initial-scale=1.0">',
                                   checkpoint_id='mobile_friendly'))
            if 'user-scalable=no' in content or 'maximum-scale=1' in content:
                issues.append(_iss('html', 'info', 'Viewport disables zoom',
                                   'Disabling zoom is bad for accessibility.',
                                   'Remove user-scalable=no / maximum-scale=1',
                                   checkpoint_id='mobile_friendly'))

        # Look for media queries in any inline <style> blocks
        has_media_query = '@media' in html

        # Look for common responsive frameworks
        responsive_hints = ['bootstrap', 'tailwind', 'foundation', 'bulma', 'mui',
                            'flexbox', 'grid-template']
        has_responsive_lib = any(h in html.lower() for h in responsive_hints)

        if not has_media_query and not has_responsive_lib:
            issues.append(_iss('html', 'warning',
                               'No responsive CSS detected',
                               'No @media queries or responsive framework found.',
                               'Add @media (max-width: 768px) breakpoints, '
                               'or use Bootstrap / Tailwind for mobile layouts',
                               checkpoint_id='mobile_friendly'))

        return issues

    # ---------- Checklist #27: Inquiry Form ----------
    def _check_inquiry_form(self, soup) -> list[dict]:
        """
        Verify a contact/inquiry form exists with:
          - An action attribute
          - Email input
          - Submit button
          - Reasonable label/name fields
        We CANNOT verify it actually delivers — that requires submitting it.
        """
        issues: list[dict] = []
        forms = soup.find_all('form')

        if not forms:
            issues.append(_iss('content', 'warning', 'No inquiry form on page',
                               'Buyers cannot easily request a quote or contact you.',
                               'Add a contact form with name, email, message, submit',
                               checkpoint_id='inquiry_form'))
            return issues

        # Check at least one form looks like a real inquiry form
        good_form = False
        for f in forms:
            inputs = f.find_all(['input', 'textarea'])
            input_types = [i.get('type', '').lower() for i in inputs]
            input_names = ' '.join((i.get('name') or '') + ' ' + (i.get('placeholder') or '')
                                   for i in inputs).lower()

            has_email = 'email' in input_types or 'email' in input_names
            has_submit = bool(f.find(['button', 'input'], type='submit')) or \
                         any('submit' in (i.get('type') or '').lower() for i in inputs)
            has_action = bool(f.get('action'))

            if has_email and (has_submit or has_action):
                good_form = True
                if not has_action:
                    issues.append(_iss('html', 'info', 'Inquiry form has no action',
                                       'Form may submit via JS — manually verify it sends emails.',
                                       'Add action="/submit" or test the JS submit handler',
                                       checkpoint_id='inquiry_form'))
                break

        if not good_form:
            issues.append(_iss('content', 'warning', 'Inquiry form looks incomplete',
                               'Found form(s) but none with email field + submit.',
                               'Ensure form has: name, email, message, submit button',
                               checkpoint_id='inquiry_form'))

        return issues

    # ---------- 360-degree: security headers ----------
    def _check_security_headers(self, page) -> list[dict]:
        """
        Check HTTP response headers for security best practices.
        These are passed in via page['response_headers'] from the crawler.
        Signals professionalism and protects users from common attacks.

        Maps to the 'mobile_friendly' / technical checkpoint since there's
        no dedicated security checkpoint in the 27-point list.
        """
        issues: list[dict] = []
        # Only skip if we genuinely don't have header info (e.g. fetch failed
        # without recording headers). An empty dict means "we got the response
        # but no security headers were set" — which is exactly the issue we
        # want to flag.
        if 'response_headers' not in page:
            return issues
        headers = {k.lower(): v for k, v in (page.get('response_headers') or {}).items()}

        # Strict-Transport-Security (HSTS) — protects against downgrade attacks
        if 'strict-transport-security' not in headers:
            issues.append(_iss('html', 'info', 'HSTS header missing',
                               'No Strict-Transport-Security header — browsers may '
                               'allow HTTPS downgrade attacks.',
                               'Add: Strict-Transport-Security: max-age=31536000; includeSubDomains',
                               checkpoint_id='mobile_friendly'))

        # X-Content-Type-Options — prevents MIME sniffing exploits
        if 'x-content-type-options' not in headers:
            issues.append(_iss('html', 'info', 'X-Content-Type-Options missing',
                               'Browsers may sniff content types and execute unexpected code.',
                               'Add: X-Content-Type-Options: nosniff',
                               checkpoint_id='mobile_friendly'))

        # Content-Security-Policy — protects against XSS
        if 'content-security-policy' not in headers:
            issues.append(_iss('html', 'info', 'Content-Security-Policy missing',
                               'No CSP defined — XSS attacks have a wider surface.',
                               'Add a CSP header restricting script/style sources. '
                               'Start with: Content-Security-Policy: default-src \'self\'',
                               checkpoint_id='mobile_friendly'))

        return issues

    # ---------- 360-degree: mixed content (HTTPS page loading HTTP) ----------
    def _check_mixed_content(self, soup, url) -> list[dict]:
        """
        If the page is loaded over HTTPS but pulls resources over HTTP,
        browsers show a broken padlock — kills trust instantly.
        """
        issues: list[dict] = []
        if not url.startswith('https://'):
            return issues  # Only matters on HTTPS pages

        http_resources: list[str] = []
        for tag, attr in [('img', 'src'), ('script', 'src'),
                          ('link', 'href'), ('iframe', 'src')]:
            for el in soup.find_all(tag):
                val = el.get(attr, '') or ''
                if val.startswith('http://'):
                    http_resources.append(f'<{tag}> {val[:60]}')
                    if len(http_resources) >= 5:
                        break
            if len(http_resources) >= 5:
                break

        if http_resources:
            example = http_resources[0]
            issues.append(_iss('html', 'warning',
                               f'{len(http_resources)} resources loaded over HTTP',
                               f'HTTPS pages with HTTP resources show a broken '
                               f'padlock in browsers — looks unprofessional. '
                               f'First: {example}',
                               'Update all http:// resource URLs to https:// '
                               '(or use protocol-relative //)',
                               checkpoint_id='mobile_friendly'))
        return issues

    # ---------- 360-degree: social media completeness (Twitter Cards) ----------
    def _check_social_media_tags(self, soup) -> list[dict]:
        """
        Beyond basic Open Graph (which we already check elsewhere), verify
        Twitter Card tags are present. Important for B2B social sharing.
        """
        issues: list[dict] = []

        # OG image is the most important — without it, social previews are bare
        if not soup.find('meta', property='og:image'):
            issues.append(_iss('seo', 'info', 'og:image missing',
                               'When users share the page on LinkedIn/Twitter/'
                               'Facebook/WhatsApp, no preview image will show — '
                               'massive engagement drop.',
                               '<meta property="og:image" content="https://yoursite.com/og-image.jpg"> '
                               '(use a 1200×630 image)',
                               checkpoint_id='schema'))

        # Twitter Card type — controls what Twitter/X shows when shared
        if not soup.find('meta', attrs={'name': 'twitter:card'}):
            issues.append(_iss('seo', 'info', 'Twitter Card meta missing',
                               'Twitter/X shows a basic link preview instead of a rich card.',
                               '<meta name="twitter:card" content="summary_large_image">',
                               checkpoint_id='schema'))

        return issues

    # ============================================================
    #  END NEW CHECKLIST-DRIVEN CHECKS
    # ============================================================

    # ---------- Metals: chemical composition ----------
    def _check_chemical(self, text, grades) -> list[dict]:
        issues: list[dict] = []
        for raw in set(g.upper().replace(' ', '') for g in grades):
            clean = raw.replace('SS', '').strip()
            ref = ASTM_GRADES.get(clean)
            if not ref:
                continue

            if 'carbon' not in text.lower():
                issues.append(_iss('chemical', 'critical',
                                   f'{raw}: Carbon % missing',
                                   'Chemical composition table lacks Carbon.',
                                   f'Carbon (C): {ref.get("C", "")} max',
                                   checkpoint_id='chemical_table'))
                continue

            m = re.search(r'[Cc]arbon[^\d]*(\d+\.?\d*)', text)
            if m and ref.get('C'):
                try:
                    val = float(m.group(1))
                    rmax = float(re.findall(r'[\d.]+', ref['C'])[-1])
                    if val > rmax + 0.001:
                        issues.append(_iss('chemical', 'critical',
                                           f'{raw}: Carbon out of spec ({val}% > {rmax}% max)',
                                           'Wrong value vs ASTM specification.',
                                           f'Carbon: {ref["C"]} max',
                                           checkpoint_id='chemical_table'))
                except Exception:
                    pass

            for elem, label in [('Cr', 'Chromium'), ('Ni', 'Nickel'), ('Mo', 'Molybdenum')]:
                if ref.get(elem) and label not in text:
                    issues.append(_iss('chemical', 'warning',
                                       f'{raw}: {label} missing',
                                       f'{label} is important for this grade.',
                                       f'{label} ({elem}): {ref[elem]}',
                                       checkpoint_id='chemical_table'))

        if not grades:
            issues.append(_iss('chemical', 'info', 'No grade detected',
                               'Mention grade clearly.',
                               'e.g. SS 304, SS 316L, ASTM A312 TP304',
                               checkpoint_id='chemical_table'))
        return issues

    # ---------- Metals: mechanical properties ----------
    def _check_mechanical(self, text, grades) -> list[dict]:
        issues: list[dict] = []
        if not any(t in text.lower() for t in ['tensile', 'yield', 'elongation', 'mpa', 'hardness']):
            issues.append(_iss('mechanical', 'critical',
                               'Mechanical properties table missing',
                               'Buyers need tensile/yield/elongation data.',
                               'Add table: Tensile | Yield | Elongation | Hardness',
                               checkpoint_id='mechanical_table'))
            return issues

        for raw in set(g.upper().replace(' ', '') for g in grades):
            clean = raw.replace('SS', '').strip()
            ref = MECHANICAL_PROPS.get(clean)
            if not ref:
                continue
            for key, label in [('tensile', 'Tensile Strength'), ('yield', 'Yield Strength'),
                               ('elongation', 'Elongation'), ('hardness', 'Hardness')]:
                if key not in text.lower():
                    issues.append(_iss('mechanical', 'warning',
                                       f'{raw}: {label} missing',
                                       f'Mechanical table lacks {label}.',
                                       f'{label}: {ref.get(key, "—")} (per ASTM)',
                                       checkpoint_id='mechanical_table'))
        return issues

    # ---------- Metals: ASTM standards ----------
    def _check_astm_standards(self, text, url) -> list[dict]:
        issues: list[dict] = []
        ul = url.lower()
        product_type = next(
            (p for p in ASTM_STANDARDS if p in ul or p in text.lower()),
            None,
        )
        if product_type:
            missing = [s for s in ASTM_STANDARDS[product_type]
                       if s.replace(' ', '') not in text.replace(' ', '')]
            if missing:
                issues.append(_iss('astm', 'warning',
                                   f'Missing standards: {", ".join(missing[:3])}',
                                   f'Required for {product_type} pages.',
                                   f'Add: {" | ".join(missing)}',
                                   checkpoint_id='specifications'))
        if not any(s in text for s in ['ASTM', 'ASME', 'DIN', 'EN ', 'JIS']):
            issues.append(_iss('astm', 'critical', 'No standards mentioned',
                               'No ASTM/ASME/DIN/EN/JIS at all.',
                               'Add: ASTM A312 / ASME SA312 / EN 10217-7',
                               checkpoint_id='specifications'))
        return issues

    # ---------- Metals: equivalent grades ----------
    def _check_equivalents(self, text, grades) -> list[dict]:
        issues: list[dict] = []
        for raw in set(g.upper().replace(' ', '') for g in grades):
            clean = raw.replace('SS', '').strip()
            ref = EQUIVALENT_GRADES.get(clean)
            if not ref:
                continue
            missing = next((k for k, v in ref.items() if v and v not in text), None)
            if missing:
                issues.append(_iss('equivalent', 'info',
                                   f'{raw}: {missing} equivalent missing',
                                   'International buyers need equivalents.',
                                   f'UNS {ref.get("UNS", "")} | EN {ref.get("EN", "")} | '
                                   f'DIN {ref.get("DIN", "")} | JIS {ref.get("JIS", "")}',
                                   checkpoint_id='equivalent_table'))
        return issues

    # ---------- E-commerce ----------
    def _check_ecommerce(self, soup, text) -> dict:
        issues: list[dict] = []
        score = 100
        if not re.search(r'₹|Rs\.?|\$|€|£|price|cost', text, re.I):
            issues.append(_iss('ecommerce', 'warning', 'Price missing',
                               'Product page without visible price.',
                               'Show price or "Get Quote" button'))
            score -= 20
        if not soup.find('script', type='application/ld+json'):
            issues.append(_iss('ecommerce', 'warning', 'Product schema missing',
                               'Will not appear in Google Shopping.',
                               '{"@type":"Product","name":"...","offers":{"price":"..."}}'))
            score -= 15
        if len(soup.find_all('img')) < 2:
            issues.append(_iss('ecommerce', 'warning', 'Few product images',
                               'Need 3-5+ images per product.',
                               'Add multiple angles, zoom view'))
            score -= 10
        if 'review' not in text.lower():
            issues.append(_iss('ecommerce', 'info', 'Reviews missing',
                               'No social proof.',
                               'Add customer reviews/ratings section'))
            score -= 8
        return {'i': issues, 's': max(0, score)}

    # ---------- SaaS ----------
    def _check_saas(self, soup, text) -> dict:
        issues: list[dict] = []
        score = 100
        if 'pricing' not in text.lower():
            issues.append(_iss('saas', 'warning', 'No pricing info',
                               'Transparent pricing builds trust.',
                               'Add a pricing page link or table'))
            score -= 20
        if not any(t in text.lower() for t in ['testimonial', 'review', 'customer', 'client']):
            issues.append(_iss('saas', 'warning', 'No social proof',
                               'No testimonials visible.',
                               'Add testimonials, client logos, ratings'))
            score -= 15
        if not any(t in text.lower() for t in ['free trial', 'demo', 'get started', 'signup']):
            issues.append(_iss('saas', 'warning', 'No trial/demo CTA',
                               'SaaS sites need a clear conversion path.',
                               'Add "Start Free Trial" or "Book Demo" button'))
            score -= 10
        return {'i': issues, 's': max(0, score)}

    # ---------- Healthcare ----------
    def _check_healthcare(self, soup, text) -> dict:
        issues: list[dict] = []
        score = 100
        if not re.search(r'\d{10}|\+91|\+1', text):
            issues.append(_iss('healthcare', 'critical', 'Phone number missing',
                               'No emergency contact visible.',
                               'Show phone prominently in header'))
            score -= 25
        if 'appointment' not in text.lower() and 'book' not in text.lower():
            issues.append(_iss('healthcare', 'critical', 'Appointment booking missing',
                               'No way to book online.',
                               'Add a "Book Appointment" button'))
            score -= 20
        if not soup.find('script', type='application/ld+json'):
            issues.append(_iss('healthcare', 'warning', 'Medical schema missing',
                               'Clinic info will not appear richly in Google.',
                               '{"@type":"MedicalOrganization","telephone":"..."}'))
            score -= 12
        return {'i': issues, 's': max(0, score)}

    # ---------- Real estate ----------
    def _check_realestate(self, soup, text) -> dict:
        issues: list[dict] = []
        score = 100
        if not re.search(r'₹|cr|lakh|sqft|bhk', text, re.I):
            issues.append(_iss('realestate', 'warning', 'Price/size info missing',
                               'Property details are unclear.',
                               'Add price, area (sqft), BHK clearly'))
            score -= 20
        img_count = len(soup.find_all('img'))
        if img_count < 5:
            issues.append(_iss('realestate', 'warning',
                               f'Too few property images ({img_count})',
                               'Need 8-10+ photos.',
                               'Add: exterior, interior, floor plan, amenities'))
            score -= 15
        if 'map' not in text.lower() and 'location' not in text.lower():
            issues.append(_iss('realestate', 'warning', 'Location/map missing',
                               'No Google Maps embed.',
                               'Embed Google Maps and list nearby landmarks'))
            score -= 15
        return {'i': issues, 's': max(0, score)}


# ============================================================
#  Report generators (HTML, Excel, CSV — for download)
# ============================================================

class ReportGen:
    """Generate downloadable reports in HTML, Excel, and CSV formats."""

    def _build_recommendations_html(self, recs: list[dict]) -> str:
        """Render the top recommendations as a polished HTML block for the report."""
        if not recs:
            return ''

        theme_colors = {
            'seo':     ('#2b6cb0', '#ebf4ff'),  # blue
            'design':  ('#6b46c1', '#faf5ff'),  # purple
            'content': ('#2f855a', '#f0fff4'),  # green
            'trust':   ('#c05621', '#fffaf0'),  # amber
            'tech':    ('#c53030', '#fff5f5'),  # red
        }
        sev_badge = {
            'critical': ('#742a2a', '#fed7d7', 'CRITICAL'),
            'warning':  ('#7b341e', '#feebc8', 'WARNING'),
            'info':     ('#2b6cb0', '#ebf4ff', 'IMPROVEMENT'),
        }

        parts: list[str] = [
            '<div class="recs">',
            '<h2>Top Recommendations</h2>',
            '<p class="r-hint">Design-based suggestions ordered by impact across your site. '
            'Start at the top — those have the highest leverage.</p>',
        ]

        for idx, r in enumerate(recs, 1):
            tc_fg, tc_bg = theme_colors.get(r['theme'], ('#4a5568', '#f7fafc'))
            sev_fg, sev_bg, sev_label = sev_badge.get(r['severity'], ('#4a5568', '#edf2f7', 'INFO'))
            actions_li = ''.join(
                f'<li>{self._html_escape(a)}</li>' for a in r['action_steps']
            )
            parts.append(
                f'<div class="r-card">'
                f'<div class="r-head">'
                f'<span class="r-num">#{idx}</span>'
                f'<span class="r-theme" style="color:{tc_fg};background:{tc_bg}">'
                f'{self._html_escape(r["theme_label"])}'
                f'</span>'
                f'<span class="r-sev" style="color:{sev_fg};background:{sev_bg}">{sev_label}</span>'
                f'<span class="r-scope">{r["affected_pages"]} of {r["total_pages"]} pages</span>'
                f'</div>'
                f'<h3 class="r-headline">{self._html_escape(r["headline"])}</h3>'
                f'<p class="r-why">{self._html_escape(r["why"])}</p>'
                f'<details class="r-actions">'
                f'<summary>Action steps ({len(r["action_steps"])})</summary>'
                f'<ol>{actions_li}</ol>'
                f'</details>'
                f'</div>'
            )

        parts.append('</div>')
        return ''.join(parts)

    def _build_checklist_html(self, checklist_summary: list[dict], total_pages: int) -> str:
        """Render the 27-point checklist summary as an HTML block."""
        if not checklist_summary or total_pages == 0:
            return ''

        groups = ['On-page SEO', 'Images', 'Content', 'Product Tables',
                  'Page Content', 'Technical']

        parts: list[str] = [
            '<div class="cls">',
            '<h2>27-Point Checklist Summary</h2>',
            f'<p class="cs-hint">Site-wide compliance across all {total_pages} '
            'crawled page(s). Each bar shows pass / fail / info / skipped / manual share.</p>',
        ]

        for group in groups:
            group_items = [c for c in checklist_summary if c['group'] == group]
            if not group_items:
                continue
            parts.append(f'<div class="cs-grp">{group}</div>')
            for cp in group_items:
                total = cp['total'] or 1
                segs = []
                for status, cls in [('pass', 'cs-pass'), ('fail', 'cs-fail'),
                                     ('info', 'cs-info'), ('skipped', 'cs-skipped'),
                                     ('manual', 'cs-manual')]:
                    pct = round(cp[status] / total * 100)
                    if pct > 0:
                        segs.append(f'<span class="cs-seg {cls}" '
                                    f'style="width:{pct}%" '
                                    f'title="{cp[status]} pages: {status}"></span>')
                bar = ''.join(segs) or '<span class="cs-seg cs-skipped" style="width:100%"></span>'

                # Counts text
                count_parts = []
                if cp['fail']:
                    count_parts.append(f'<span class="ng">{cp["fail"]} fail</span>')
                if cp['pass']:
                    count_parts.append(f'<span class="ok">{cp["pass"]} pass</span>')
                if cp['skipped']:
                    count_parts.append(f'{cp["skipped"]} skip')
                if cp['manual']:
                    count_parts.append(f'{cp["manual"]} manual')
                counts = ' · '.join(count_parts) or '—'

                parts.append(
                    f'<div class="cs-i" title="{self._html_escape(cp["description"])}">'
                    f'<span class="cs-n">{cp["number"]}</span>'
                    f'<span class="cs-l">{self._html_escape(cp["label"])}</span>'
                    f'<div class="cs-bar">{bar}</div>'
                    f'<span class="cs-cnt">{counts}</span>'
                    f'</div>'
                )

        parts.append(
            '<div class="cs-leg">'
            '<span><span class="cs-dot cs-pass"></span>Pass</span>'
            '<span><span class="cs-dot cs-fail"></span>Fail</span>'
            '<span><span class="cs-dot cs-info"></span>Info</span>'
            '<span><span class="cs-dot cs-skipped"></span>Skipped</span>'
            '<span><span class="cs-dot cs-manual"></span>Manual review</span>'
            '</div></div>'
        )
        return ''.join(parts)

    @staticmethod
    def _html_escape(s: str) -> str:
        return (str(s).replace('&', '&amp;').replace('<', '&lt;')
                .replace('>', '&gt;').replace('"', '&quot;'))

    def html(self, results: list[dict], site_url: str) -> str:
        """Return a complete standalone HTML report (string)."""
        now = datetime.now().strftime('%d %b %Y %I:%M %p')
        tp = len(results)
        ti = sum(len(r['issues']) for r in results)
        crit = sum(sum(1 for i in r['issues'] if i['severity'] == 'critical') for r in results)
        warn = sum(sum(1 for i in r['issues'] if i['severity'] == 'warning') for r in results)
        avg = round(sum(r['scores'].get('overall', 0) for r in results) / max(1, tp))

        sev_bg = {'critical': '#fff0f0', 'warning': '#fffbe6', 'info': '#f0f7ff'}
        sev_bd = {'critical': '#e53e3e', 'warning': '#dd6b20', 'info': '#3182ce'}
        sev_bdg = {'critical': '#c53030', 'warning': '#c05621', 'info': '#2b6cb0'}

        # ---- Build the 27-point checklist summary block ----
        cl_summary = aggregate_checklist(results)
        cl_html = self._build_checklist_html(cl_summary, tp)

        # ---- Build the Executive Summary / Recommendations block ----
        # This is the "what to actually DO" panel — design-based suggestions
        # prioritized by impact across the whole site. Shown right after the
        # score cards, before any of the per-page detail.
        recs = generate_recommendations(results)
        recs_html = self._build_recommendations_html(recs[:7])  # top 7 in report

        pages_html = ''
        for idx, r in enumerate(results):
            sc = r['scores'].get('overall', 0)
            fg, bg = score_color(sc)
            crit_c = sum(1 for i in r['issues'] if i['severity'] == 'critical')
            warn_c = sum(1 for i in r['issues'] if i['severity'] == 'warning')
            info_c = sum(1 for i in r['issues'] if i['severity'] == 'info')

            bars = ''.join(
                f'<div style="display:flex;align-items:center;gap:6px;margin:3px 0">'
                f'<span style="width:95px;font-size:10px;color:#718096;'
                f'text-transform:uppercase;letter-spacing:.3px">{cat}</span>'
                f'<div style="flex:1;height:5px;background:#e2e8f0;border-radius:3px">'
                f'<div style="width:{v}%;height:100%;background:{score_color(v)[0]};'
                f'border-radius:3px"></div></div>'
                f'<b style="width:26px;text-align:right;font-size:11px;'
                f'color:{score_color(v)[0]}">{v}</b></div>'
                for cat, v in r['scores'].items() if cat != 'overall'
            )

            iss_html = ''
            for i in sorted(r['issues'], key=lambda x: SEV_ORDER.get(x['severity'], 9)):
                fix_html = ''
                if i.get('fix'):
                    fe = (i['fix'].replace('&', '&amp;')
                                  .replace('<', '&lt;').replace('>', '&gt;'))
                    fix_html = (
                        f'<div style="background:#f7fafc;border-radius:4px;padding:7px 10px;'
                        f'margin-top:5px"><span style="font-size:9px;font-weight:700;'
                        f'color:#38a169;letter-spacing:.4px">FIX</span>'
                        f'<pre style="font-size:11px;margin:3px 0 0;white-space:pre-wrap;'
                        f'word-break:break-all;font-family:monospace;color:#2d3748">{fe}</pre></div>'
                    )
                iss_html += (
                    f'<div style="padding:9px 11px;border-radius:5px;'
                    f'background:{sev_bg.get(i["severity"], "#fff")};'
                    f'border-left:3px solid {sev_bd.get(i["severity"], "#ccc")};'
                    f'margin-bottom:5px"><div style="display:flex;align-items:flex-start;'
                    f'gap:6px;margin-bottom:4px">'
                    f'<span style="background:#edf2f7;color:#4a5568;font-size:9px;'
                    f'font-weight:700;padding:2px 5px;border-radius:8px;'
                    f'white-space:nowrap">{i["category"].upper()}</span>'
                    f'<span style="background:{sev_bdg.get(i["severity"], "#666")};'
                    f'color:#fff;font-size:9px;font-weight:700;padding:2px 5px;'
                    f'border-radius:8px;white-space:nowrap">{i["severity"].upper()}</span>'
                    f'<b style="font-size:12px;color:#2d3748">{i["title"]}</b></div>'
                    f'<p style="font-size:12px;color:#718096;line-height:1.4;'
                    f'margin:0">{i["description"]}</p>{fix_html}</div>'
                )

            # Pre-compute the no-issues fallback HTML outside the f-string.
            # Python 3.11 does NOT allow backslashes inside f-string {...} blocks
            # (PEP 701 lifted that in 3.12 but Render runs 3.11). This is the
            # exact line that crashed our deploy — keep it here, do NOT inline.
            no_issues_html = '<p style="color:#a0aec0;font-size:13px">No issues found.</p>'
            issues_block = iss_html or no_issues_html

            pages_html += (
                f'<div class="pc" id="pc{idx}" data-score="{sc}" data-crit="{crit_c}">'
                f'<div class="ph" onclick="t({idx})">'
                f'<div style="width:40px;height:40px;border-radius:50%;background:{bg};'
                f'color:{fg};font-weight:700;font-size:13px;display:flex;align-items:center;'
                f'justify-content:center;flex-shrink:0">{sc}</div>'
                f'<div style="flex:1;min-width:0">'
                f'<div style="font-weight:600;font-size:13px;overflow:hidden;'
                f'text-overflow:ellipsis;white-space:nowrap">{r["title"] or "(No Title)"}</div>'
                f'<div style="font-size:11px;color:#3182ce;overflow:hidden;'
                f'text-overflow:ellipsis;white-space:nowrap">{r["url"]}</div>'
                f'<div style="display:flex;gap:8px;margin-top:3px;font-size:11px;'
                f'color:#718096;flex-wrap:wrap">'
                f'<span style="color:#c53030;font-weight:600">{crit_c} Critical</span>'
                f'<span style="color:#c05621;font-weight:600">{warn_c} Warn</span>'
                f'<span>{info_c} Info</span>'
                f'<span>{r["word_count"]}w</span>'
                f'<span>{r["response_time"]:.2f}s</span>'
                f'<span style="background:#ebf4ff;color:#2b6cb0;padding:1px 5px;'
                f'border-radius:3px;font-size:10px">{r.get("industry", "?").upper()}</span>'
                f'</div></div>'
                f'<div style="font-size:18px;color:#a0aec0" id="ti{idx}">▾</div></div>'
                f'<div id="pb{idx}" style="display:none;border-top:1px solid #f0f0f0;'
                f'padding:12px 14px"><div style="margin-bottom:10px">{bars}</div>'
                f'{issues_block}</div>'
                f'</div>'
            )

        fgav, _ = score_color(avg)
        return (
            '<!DOCTYPE html><html><head><meta charset="UTF-8">'
            '<meta name="viewport" content="width=device-width,initial-scale=1">'
            f'<title>Audit — {site_url}</title>'
            '<style>*{box-sizing:border-box;margin:0;padding:0}'
            'body{font-family:system-ui,sans-serif;background:#f7fafc;color:#2d3748;font-size:14px}'
            '.bar{background:#1a202c;color:#fff;padding:12px 18px}'
            '.bar h1{font-size:16px;font-weight:600}'
            '.bar small{color:#a0aec0;font-size:11px}'
            '.stats{display:grid;grid-template-columns:repeat(5,1fr);gap:8px;padding:14px 18px}'
            '.stat{background:#fff;border-radius:8px;padding:12px;text-align:center;'
            'box-shadow:0 1px 2px rgba(0,0,0,.06)}'
            '.stat .n{font-size:24px;font-weight:700}'
            '.stat .l{font-size:10px;color:#718096;text-transform:uppercase;'
            'letter-spacing:.4px;margin-top:2px}'
            '.tb{padding:0 18px 10px;display:flex;gap:6px;flex-wrap:wrap;align-items:center}'
            '.fb{padding:5px 11px;border-radius:14px;border:1.5px solid #e2e8f0;'
            'background:#fff;cursor:pointer;font-size:11px;font-weight:500;color:#4a5568}'
            '.fb.on{border-color:#3182ce;background:#ebf8ff;color:#2b6cb0}'
            '.sr{flex:1;min-width:160px;padding:5px 11px;border:1.5px solid #e2e8f0;'
            'border-radius:14px;font-size:12px}'
            '.pg{padding:0 18px 24px;display:flex;flex-direction:column;gap:7px}'
            '.pc{background:#fff;border-radius:9px;box-shadow:0 1px 3px rgba(0,0,0,.06);overflow:hidden}'
            '.ph{display:flex;align-items:center;gap:10px;padding:11px 13px;cursor:pointer}'
            '.ph:hover{background:#f7fafc}'
            '@media(max-width:600px){.stats{grid-template-columns:repeat(2,1fr)}}'
            # Recommendations / Executive Summary block
            '.recs{margin:6px 18px 18px;background:#fff;border-radius:9px;'
            'box-shadow:0 1px 3px rgba(0,0,0,.06);padding:14px 16px}'
            '.recs h2{font-size:15px;font-weight:600;margin-bottom:4px}'
            '.recs .r-hint{font-size:11.5px;color:#718096;margin-bottom:14px}'
            '.recs .r-card{border:1px solid #e2e8f0;border-radius:7px;'
            'padding:12px 14px;margin-bottom:9px;background:#fafbfc}'
            '.recs .r-head{display:flex;align-items:center;gap:8px;'
            'flex-wrap:wrap;margin-bottom:6px;font-size:11px}'
            '.recs .r-num{font-weight:700;color:#a0aec0;font-size:12px}'
            '.recs .r-theme,.recs .r-sev{padding:2px 7px;border-radius:3px;'
            'font-weight:600;font-size:10px;letter-spacing:.4px}'
            '.recs .r-scope{margin-left:auto;color:#718096;font-size:11px}'
            '.recs .r-headline{font-size:14px;font-weight:600;color:#2d3748;'
            'margin:2px 0 6px}'
            '.recs .r-why{font-size:12.5px;color:#4a5568;line-height:1.55;margin:0 0 8px}'
            '.recs .r-actions{font-size:12px;color:#2d3748}'
            '.recs .r-actions summary{cursor:pointer;font-weight:600;color:#3182ce;'
            'padding:4px 0;user-select:none}'
            '.recs .r-actions summary:hover{color:#2c5282}'
            '.recs .r-actions ol{margin:6px 0 0 22px;line-height:1.6}'
            '.recs .r-actions li{margin-bottom:3px}'
            # Existing checklist styles
            '.cls{margin:6px 18px 18px;background:#fff;border-radius:9px;'
            'box-shadow:0 1px 3px rgba(0,0,0,.06);padding:14px 16px}'
            '.cls h2{font-size:14px;font-weight:600;margin-bottom:4px}'
            '.cls .cs-hint{font-size:11px;color:#718096;margin-bottom:10px}'
            '.cls .cs-grp{font-size:10px;font-weight:700;text-transform:uppercase;'
            'letter-spacing:.4px;color:#4a5568;margin:14px 0 5px;'
            'padding-bottom:3px;border-bottom:1px solid #e2e8f0}'
            '.cls .cs-grp:first-of-type{margin-top:2px}'
            '.cls .cs-i{display:grid;grid-template-columns:24px 1fr 140px 80px;'
            'align-items:center;gap:10px;padding:5px 6px;border-radius:5px;font-size:12px}'
            '.cls .cs-i:hover{background:#f7fafc}'
            '.cls .cs-n{display:inline-flex;align-items:center;justify-content:center;'
            'width:20px;height:20px;border-radius:50%;background:#e2e8f0;color:#4a5568;'
            'font-weight:700;font-size:10px}'
            '.cls .cs-l{font-weight:500;color:#2d3748}'
            '.cls .cs-bar{display:flex;height:7px;border-radius:4px;overflow:hidden;background:#edf2f7}'
            '.cls .cs-seg{display:block;height:100%}'
            '.cls .cs-pass{background:#48bb78}'
            '.cls .cs-fail{background:#f56565}'
            '.cls .cs-info{background:#4299e1}'
            '.cls .cs-skipped{background:#a0aec0}'
            '.cls .cs-manual{background:#ed8936}'
            '.cls .cs-cnt{font-size:10.5px;color:#718096;text-align:right}'
            '.cls .cs-cnt .ok{color:#2f855a;font-weight:600}'
            '.cls .cs-cnt .ng{color:#c53030;font-weight:600}'
            '.cls .cs-leg{display:flex;flex-wrap:wrap;gap:14px;margin-top:14px;'
            'padding-top:10px;border-top:1px solid #e2e8f0;font-size:10.5px;color:#4a5568}'
            '.cls .cs-leg span{display:inline-flex;align-items:center;gap:5px}'
            '.cls .cs-dot{display:inline-block;width:8px;height:8px;border-radius:50%}'
            '@media(max-width:600px){.cls .cs-i{grid-template-columns:22px 1fr;'
            'grid-template-areas:"n l" "bar bar" "cnt cnt";gap:6px}'
            '.cls .cs-n{grid-area:n}.cls .cs-l{grid-area:l}'
            '.cls .cs-bar{grid-area:bar}.cls .cs-cnt{grid-area:cnt;text-align:left}}'
            '</style></head><body>'
            f'<div class="bar"><h1>Website Audit Report</h1>'
            f'<small>{site_url} &nbsp;·&nbsp; {now} &nbsp;·&nbsp; {tp} pages</small></div>'
            f'<div class="stats">'
            f'<div class="stat"><div class="n">{tp}</div><div class="l">Pages</div></div>'
            f'<div class="stat"><div class="n" style="color:#c53030">{crit}</div>'
            f'<div class="l">Critical</div></div>'
            f'<div class="stat"><div class="n" style="color:#c05621">{warn}</div>'
            f'<div class="l">Warnings</div></div>'
            f'<div class="stat"><div class="n">{ti}</div><div class="l">Total Issues</div></div>'
            f'<div class="stat"><div class="n" style="color:{fgav}">{avg}</div>'
            f'<div class="l">Avg Score</div></div></div>'
            f'<div class="tb">'
            f'<button class="fb on" onclick="f(\'all\',this)">All ({tp})</button>'
            f'<button class="fb" onclick="f(\'crit\',this)">Critical Only</button>'
            f'<button class="fb" onclick="f(\'low\',this)">Score &lt;60</button>'
            f'<button class="fb" onclick="f(\'good\',this)">Score &ge;80</button>'
            f'<input class="sr" placeholder="Search URL / title..." oninput="s(this.value)"></div>'
            f'{recs_html}'
            f'{cl_html}'
            f'<div class="pg" id="pg">{pages_html}</div>'
            '<script>'
            'function t(i){var b=document.getElementById("pb"+i),'
            'tc=document.getElementById("ti"+i);'
            'if(b.style.display==="none"){b.style.display="block";'
            'tc.style.transform="rotate(180deg)"}else{b.style.display="none";'
            'tc.style.transform=""}}'
            'function f(type,btn){document.querySelectorAll(".fb").forEach(b=>'
            'b.classList.remove("on"));btn.classList.add("on");'
            'document.querySelectorAll(".pc").forEach(c=>{var sc=+c.dataset.score,'
            'cr=+c.dataset.crit;c.style.display=(type==="all"||(type==="crit"&&cr>0)||'
            '(type==="low"&&sc<60)||(type==="good"&&sc>=80))?"":"none"})}'
            'function s(q){q=q.toLowerCase();document.querySelectorAll(".pc").forEach(c=>{'
            'c.style.display=c.querySelector(".ph").textContent.toLowerCase().includes(q)?"":"none"})}'
            '</script></body></html>'
        )

    def excel(self, results: list[dict]) -> bytes:
        """Return the Excel report as bytes (so Flask can stream it)."""
        wb = openpyxl.Workbook()
        hf = PatternFill('solid', fgColor='1A202C')
        hft = Font(color='FFFFFF', bold=True, size=10)

        ws = wb.active
        ws.title = 'Summary'
        ws.append(['URL', 'Title', 'Industry', 'Score', 'SEO', 'HTML', 'Perf', 'Content',
                   'Chemical', 'Mech', 'Standards', 'Equiv', 'Critical', 'Warnings',
                   'Words', 'Load(s)'])
        for c in ws[1]:
            c.fill, c.font = hf, hft

        for r in results:
            sc = r['scores']
            ws.append([
                r['url'], r['title'], r.get('industry', ''),
                sc.get('overall', 0), sc.get('seo', 0), sc.get('html', 0),
                sc.get('performance', 0), sc.get('content', 0),
                sc.get('chemical', '-'), sc.get('mechanical', '-'),
                sc.get('standards', '-'), sc.get('equivalent', '-'),
                sum(1 for i in r['issues'] if i['severity'] == 'critical'),
                sum(1 for i in r['issues'] if i['severity'] == 'warning'),
                r['word_count'], round(r['response_time'], 2),
            ])

        for row in ws.iter_rows(min_row=2):
            for ci in [3, 4, 5, 6, 7, 8, 9, 10, 11]:
                cell = row[ci]
                if isinstance(cell.value, (int, float)):
                    v = cell.value
                    fc = '1D9E75' if v >= 80 else ('BA7517' if v >= 55 else 'E24B4A')
                    cell.font = Font(color=fc, bold=True)

        # All Issues sheet
        wi = wb.create_sheet('All Issues')
        wi.append(['URL', 'Title', 'Category', 'Severity', 'Issue', 'Description', 'Fix'])
        for c in wi[1]:
            c.fill, c.font = hf, hft

        sfills = {
            'critical': PatternFill('solid', fgColor='FEE2E2'),
            'warning':  PatternFill('solid', fgColor='FEF3C7'),
            'info':     PatternFill('solid', fgColor='DBEAFE'),
        }
        for r in results:
            for i in sorted(r['issues'], key=lambda x: SEV_ORDER.get(x['severity'], 9)):
                wi.append([r['url'], r['title'], i['category'], i['severity'],
                           i['title'], i['description'], i.get('fix', '')])
                for c in wi[wi.max_row]:
                    c.fill = sfills.get(i['severity'], PatternFill())

        # Chemical reference sheet
        wc = wb.create_sheet('Chemical Reference')
        wc.append(['Grade', 'C', 'Mn', 'Si', 'P', 'S', 'Cr', 'Ni', 'Mo', 'N/Ti'])
        for c in wc[1]:
            c.fill, c.font = hf, hft
        for g, p in ASTM_GRADES.items():
            wc.append([g, p.get('C', ''), p.get('Mn', ''), p.get('Si', ''),
                       p.get('P', ''), p.get('S', ''), p.get('Cr', ''),
                       p.get('Ni', ''), p.get('Mo', ''), p.get('N', p.get('Ti', ''))])

        # Equivalent grades sheet
        we = wb.create_sheet('Equivalent Grades')
        we.append(['Grade', 'UNS', 'EN', 'DIN', 'JIS', 'BS'])
        for c in we[1]:
            c.fill, c.font = hf, hft
        for g, eq in EQUIVALENT_GRADES.items():
            we.append([g, eq.get('UNS', ''), eq.get('EN', ''),
                       eq.get('DIN', ''), eq.get('JIS', ''), eq.get('BS', '')])

        # ---- 27-Point Checklist sheet ----
        # Rows = pages, columns = each of the 27 checkpoints + the URL.
        # Cells get a status letter (P/F/I/M/S) plus colour-coded fill so the
        # whole sheet reads at a glance.
        wcl = wb.create_sheet('27-Point Checklist')
        header = ['URL', 'Title'] + [f"{cp['number']}. {cp['label']}" for cp in CHECKLIST]
        wcl.append(header)
        for c in wcl[1]:
            c.fill, c.font = hf, hft

        status_fills = {
            'pass':    PatternFill('solid', fgColor='C6F6D5'),  # green
            'fail':    PatternFill('solid', fgColor='FED7D7'),  # red
            'info':    PatternFill('solid', fgColor='DBEAFE'),  # blue
            'manual':  PatternFill('solid', fgColor='FEEBC8'),  # amber
            'skipped': PatternFill('solid', fgColor='EDF2F7'),  # gray
        }
        status_letters = {
            'pass': 'P', 'fail': 'F', 'info': 'I',
            'manual': 'M', 'skipped': 'S',
        }
        for r in results:
            row = [r['url'], r.get('title', '')]
            checklist = {c['id']: c for c in r.get('checklist', [])}
            for cp in CHECKLIST:
                entry = checklist.get(cp['id'])
                row.append(status_letters.get(entry['status'], '-') if entry else '-')
            wcl.append(row)
            # Apply colour to the 27 checklist cells in this row
            row_idx = wcl.max_row
            for col_offset, cp in enumerate(CHECKLIST, start=3):  # cols 3..29
                entry = checklist.get(cp['id'])
                if entry:
                    cell = wcl.cell(row=row_idx, column=col_offset)
                    cell.fill = status_fills.get(entry['status'], PatternFill())
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                    cell.font = Font(bold=True, size=10)

        # Add a legend at the bottom
        wcl.append([])
        wcl.append(['Legend:', 'P = Pass', 'F = Fail', 'I = Info (minor)',
                    'M = Manual review needed', 'S = Skipped (turn on Deep Audit / target keyword)'])

        # Auto-size columns
        for ws2 in wb.worksheets:
            for col in ws2.columns:
                width = min(max(len(str(c.value or '')) for c in col) + 3, 55)
                ws2.column_dimensions[col[0].column_letter].width = width

        # Narrow the 27 checklist columns for readability
        for col_offset in range(3, 3 + len(CHECKLIST)):
            wcl.column_dimensions[wcl.cell(row=1, column=col_offset).column_letter].width = 6

        # Save to bytes buffer
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def csv_report(self, results: list[dict]) -> str:
        """Return a CSV summary as a string."""
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(['URL', 'Title', 'Industry', 'Score', 'SEO', 'HTML',
                    'Performance', 'Content', 'Critical', 'Warnings', 'Words', 'Load(s)'])
        for r in results:
            sc = r['scores']
            w.writerow([
                r['url'], r['title'], r.get('industry', ''),
                sc.get('overall', 0), sc.get('seo', 0), sc.get('html', 0),
                sc.get('performance', 0), sc.get('content', 0),
                sum(1 for i in r['issues'] if i['severity'] == 'critical'),
                sum(1 for i in r['issues'] if i['severity'] == 'warning'),
                r['word_count'], round(r['response_time'], 2),
            ])
        return buf.getvalue()
